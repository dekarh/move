from string import digits
from collections import OrderedDict
import requests, json
import openpyxl, sys
from openpyxl import Workbook

from lib import read_config, l, s, fine_phone, format_phone

wb_out = openpyxl.Workbook(write_only=True)
ws_out = wb_out.create_sheet('Лист 1')

wb_in = openpyxl.load_workbook(sys.argv[1], read_only=True)
ws_in = wb_in[wb_in.sheetnames[0]]
for i, row in enumerate(ws_in.rows):
    tek_row = []
    for j, cell in enumerate(row):
        if j == 6 and i and cell.value:
            words = cell.value.split()
            has_home = False
            for k, word in enumerate(words):
                if k and not has_home and l(word) and word.strip()[-2:].lower() != '-я' and word.strip()[-2:].lower() != '-й':
                    words[k] = 'д. ' + words[k]
                    has_home = True
            tek_row.append('Москва ' + ' '.join(words))
        else:
            tek_row.append(cell.value)
    ws_out.append(tek_row)
wb_out.save('out_' + sys.argv[1])
