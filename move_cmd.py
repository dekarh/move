# Запуск из командной строки (генерируется в move.py)

import openpyxl
import sys
import argparse
from datetime import datetime
from mysql.connector import MySQLConnection

from lib import read_config, l, s

IN_IDS = ['ID','ИД_КЛИЕНТА','CLIENT_ID']
IN_SNILS = ['СНИЛС', 'СТРАХОВОЙ_НОМЕР', 'СТРАХОВОЙ НОМЕР', 'NUMBER','СТРАХОВОЙНОМЕР']
IN_NAMES = ['ID', 'СНИЛС', 'СТРАХОВОЙ_НОМЕР', 'СТРАХОВОЙ НОМЕР', 'NUMBER', 'ФАМИЛИЯ', 'ИМЯ', 'ОТЧЕСТВО', 'ФИО']
DIR4MOVE = '/home/lekarh/Move/'


def createParser ():
    parser = argparse.ArgumentParser()
    parser.add_argument ('-sheetName')
    parser.add_argument ('-fond', default=0, type=int)
    parser.add_argument ('-agent', default=0, type=int)
    parser.add_argument ('-signer', default=0, type=int)
    parser.add_argument ('-clientOnly', default=False, type=bool)
    parser.add_argument ('-socium', default=False, type=bool)
    parser.add_argument ('-suff', default='', type=str)
    parser.add_argument ('-ourStat', default=False, type=bool)
    parser.add_argument ('-fondStat', default=False, type=bool)
    parser.add_argument ('-arhivON', default=False, type=bool)
    parser.add_argument ('-arhivOFF', default=False, type=bool)
    parser.add_argument ('-noDubPhonePartner', default=False, type=bool)


class my(object):
    def __init__(self):
        parser = createParser()
        self.args = parser.parse_args(sys.argv[2:])
        self.dbconfig = read_config(filename='move.ini', section='mysql')
        self.leSQLcl = ''
        self.leSQLco = ''

    def one(self):
        # ФИО агента
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        if self.args.agent:
            sql = "SELECT CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, '-', position_id), code " \
                  "FROM offices_staff WHERE code = %s"
            cursor.execute(sql, (self.args.agent,))
            rows = cursor.fetchall()
            agents = []
            self.agent_names = {}
            self.agent_ids = []
            for i, row in enumerate(rows):
                agents.append(row[0])
                self.agent_names[row[1]] = row[0]
                self.agent_ids.append(row[1])
        # Список названий фондов
        cursor = dbconn.cursor()
        if self.args.fond:
            sql = "SELECT CONCAT_WS(' ', id, '-', name), id FROM subdomains WHERE id IN (2,6,8,11,12,13,14)"
            cursor.execute(sql)
        rows = cursor.fetchall()
        fonds = []
        self.fond_ids = []
        self.fond_names = {}
        self.fonds_str = '1'
        for i, row in enumerate(rows):
            fonds.append(row[0])
            self.fond_names[row[1]] = row[0]
            self.fond_ids.append(row[1])
            if i == 0:
                self.fonds_str = str(row[1])
                continue
            self.fonds_str += ','+ str(row[1])
        # ФИО подписанта
        cursor = dbconn.cursor()
        if self.args.signer:
            sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                  "FROM signers WHERE id = %s"
            cursor.execute(sql, (self.args.signer,))
            rows = cursor.fetchall()
            self.signer_ids = []
            self.signer_names = {}
            signers = []
            for i, row in enumerate(rows):
                self.signer_names[row[1]] = row[0]
                self.signer_ids.append(row[1])
                signers.append(row[0])

        wb = openpyxl.load_workbook(sys.argv[1])
        if not self.args.sheetName:
            self.args.sheetName = wb.sheetnames[0]
        self.sheet = wb[self.args.sheetName]
        if not self.sheet.max_row:
            print('Файл Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
            return
        keys = {}
        last_cell = 0
        for j, row in enumerate(self.sheet.rows):
            if j == 0:
                for k, cell in enumerate(row):  # Проверяем, чтобы был СНИЛС
                    if str(cell.value).upper() in IN_IDS:
                        keys[IN_IDS[0]] = k
                if len(keys) > 0:
                    for k, cell in enumerate(row):
                        for n, name in enumerate(IN_NAMES):
                            if n == 0:
                                continue
                            if cell.value != None:
                                if str(cell.value).strip() != '':
                                    last_cell = k
                                    if str(cell.value).upper() == name:
                                        keys[name] = k

                else:
                    print('В файле ' + sys.argv[1] + ' на вкладке ' + self.args.sheetName + ' отсутствует колонка с ID')
                    return
            elif j == 1:
                for k, cell in enumerate(row):
                    if cell.value != None:
                        if str(cell.value).strip() != '':
                            if k > last_cell:
                                last_cell = k

        self.clients_ids = []
        for j, row in enumerate(self.sheet.rows):
            if j == 0:
                continue
            for k, key in enumerate(keys):
                if k == 0:
                    self.clients_ids.append(row[keys[key]].value)

        self.table = []
        for i, row in enumerate(self.sheet.rows):
            table_row = []
            for j, cell in enumerate(row):
                if j > last_cell:
                    break
                table_row.append(cell.value)
            self.table.append(table_row)

        wb_log = openpyxl.Workbook(write_only=True)

        ws_log = wb_log.create_sheet('Лог')
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Начинаем'])

        log_name = DIR4MOVE + datetime.now().strftime('%Y-%m-%d_%H-%M')
        if self.args.fond:
            log_name += 'ф' + str(self.args.fond)
        if self.args.agent:
            log_name += 'а' + str(self.args.agent)
        log_name += '.xlsx'

        if not self.args.noBackup:
            # Проверка на дубли clients
            all_clients_ids = "'" + self.clients_ids[0] + "'"
            for i, client_id in enumerate(self.clients_ids):
                if i == 0:
                    continue
                all_clients_ids += ",'" + client_id + "'"
            sql = "SELECT cl.client_id FROM clients AS cl WHERE cl.client_id IN (" + all_clients_ids + \
                  ") GROUP BY cl.client_id HAVING COUNT(cl.client_id) > 1 ORDER BY cl.client_id DESC"
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
            exit_because_doubles = False
            if len(rows) > 0:
                exit_because_doubles = True
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в clients'])
                ws_clients = wb_log.create_sheet('Дубли в clients')
                for row in rows:
                    ws_clients.append([row[0]])
            else:
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' В clients нет дублей'])
            # Проверка на дубли contracts
            sql = "SELECT co.client_id FROM contracts AS co WHERE co.client_id IN (" + all_clients_ids + \
                  ") GROUP BY co.client_id HAVING COUNT(co.client_id) > 1 ORDER BY co.client_id DESC"
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
            if len(rows) > 0:
                exit_because_doubles = True
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в contracts'])
                ws_contracts = wb_log.create_sheet('Дубли в contracts')
                for row in rows:
                    ws_contracts.append([row[0]])
            else:
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' В contracts нет дублей'])
            if exit_because_doubles:  # Если дубли в clients или contracts - ничего не переносим
                ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Аварийное завершение - дублирование записей'])
                return

        # Проверка на дубли исходной таблицы
        doubles_in_input = list(set([x for x in self.clients_ids if self.clients_ids.count(x) > 1]))
        if len(doubles_in_input) > 0:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в исходной таблице'])
            ws_input_doubles = wb_log.create_sheet('Дубли в исходной таблице')
            for row in doubles_in_input:
                ws_input_doubles.append(row)
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), ' В исходной таблице нет дублей'])

        if not self.args.noBackup:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Дублируем исходную excel таблицу в этот файл'])
            ws_input = wb_log.create_sheet('Исходная таблица')
            for table_row in self.table:
                row = []
                for cell in table_row:
                    row.append(cell)
                ws_input.append(row)
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Бэкап исходного состояния БД создан'])
            ws_backup = wb_log.create_sheet('бэкап БД')
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            sql = "SELECT cl.*, co.* FROM clients AS cl LEFT JOIN contracts AS co " \
                  "ON (cl.client_id = co.client_id) WHERE cl.client_id IN (" + all_clients_ids + ")"
            cursor.execute(sql)
            dbrows = cursor.fetchall()
            ws_backup.append(cursor.column_names)
            for dbrow in dbrows:
                row = []
                for dbcell in dbrow:
                    row.append(dbcell)
                ws_backup.append(row)

        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Состояние программы:'])
        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'файл ', sys.argv[1]])
        if self.args.fond:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', self.fond_names[self.args.fond]])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', 'не выбран'])
        if self.args.agent:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', self.agent_names[self.args.agent]])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', 'не выбран'])
        if self.args.signer:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', self.signer_names[self.args.signer]])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', 'не выбран'])
        if self.args.clientOnly:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Перенести только клиента', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Перенести только клиента', 'не выбрано'])
        if self.args.socium:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить номер Социума', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить номер Социума', 'не выбрано'])
        if self.args.suff:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Суффикс', self.args.suff])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Суффикс', 'не выбрано'])
        if self.args.ourStat:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить внутренние статусы', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить внутренние статусы', 'не выбрано'])
        if self.args.fondStat:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить статусы Фонда', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить статусы Фонда', 'не выбрано'])
        if self.args.arhivON:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Поставить флаг "Архивный"', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Поставить флаг "Архивный"', 'не выбрано'])
        if self.args.arhivOFF:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Убрать флаг "Архивный"', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Убрать флаг "Архивный"', 'не выбрано'])
        if self.args.noDubPhonePartner:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Без дублей телефонов у партнера', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Без дублей телефонов у партнера', 'не выбрано'])

        # Список телефонов у партнера в фонде в который переносим
        if self.args.noDubPhonePartner and self.args.agent:
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            cursor.execute('SELECT partner_code FROM offices_staff WHERE code = %s', (self.args.agent,))
            partner = cursor.fetchall()
            if self.partner != partner[0][0]:
                self.partner = partner[0][0]
                phones = []
                cursor = dbconn.cursor()
                sql_tel = 'SELECT phone_personal_mobile FROM clients AS cl LEFT JOIN offices_staff AS os ' \
                          'ON cl.inserted_user_code = os.code WHERE os.partner_code = %s'
                if self.args.fond:
                    cursor.execute(sql_tel + ' AND cl.subdomain_id = %s', (partner[0][0], self.args.fond))
                else:
                    cursor.execute(sql_tel, (partner[0][0],))
                phones_sql = cursor.fetchall()
                for i, phone_sql in enumerate(phones_sql):
                    if phone_sql[0] and phone_sql[0] not in phones:
                        phones.append(phone_sql[0])
                    # if i > 10000:
                    #    break
                self.phones = phones
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Формируем запросы:'])
        sql_cl = 'UPDATE clients AS cl SET'
        sql_co = 'UPDATE contracts AS co SET'
        if self.args.agent:
            sql_cl += ' cl.inserted_user_code = %s'
            sql_co += ' co.inserted_code = %s, co.agent_code = %s'
        if self.args.fond:
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ','
            sql_cl += ' cl.subdomain_id = %s'
        if self.args.arhivON or self.args.arhivOFF:
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ','
            sql_cl += ' cl.archived = %s'
        if self.args.signer:
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.signer_id = %s'
        if self.args.socium:
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.socium_contract_number = %s'
        if self.args.fondStat:
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.external_status_code = %s, co.external_status_secure_code = %s,' \
                      ' co.external_status_callcenter_code = %s'
        if self.args.ourStat.isChecked():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.status_code = %s, co.status_secure_code = %s, co.status_callcenter_code = %s'
        if self.args.suff.isChecked():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.partner_remote_id = %s'
        if sql_cl[len(sql_cl) - 3:] == 'SET':
            self.leSQLcl = ''
        else:
            self.leSQLcl = sql_cl + ' WHERE cl.client_id = %s'
        if sql_co[len(sql_co) - 3:] == 'SET' or self.args.clientOnly:
            self.leSQLco = ''
        else:
            self.leSQLco = sql_co + ' WHERE co.client_id = %s'

        tuples_clients = []  # Формируем переменные для запросов
        tuples_contracts = []
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        cursor_phones = dbconn.cursor()
        i_tek = 0
        if len(self.phones):
            ws_phones_doubles = wb_log.create_sheet('Дубли телефонов у партнера')
            ws_phones_doubles.append(['client_id', 'СНИЛС', 'телефон'])
        for i, client_id in enumerate(self.clients_ids):
            if len(self.phones):
                cursor_phones.execute('SELECT phone_personal_mobile, number FROM clients AS cl WHERE cl.client_id = %s',
                                      (client_id,))
                rows = cursor_phones.fetchall()
                if rows[0][0] in self.phones:
                    ws_phones_doubles.append([client_id, rows[0][1], rows[0][0]])
                    continue
            tuple_client = tuple()
            tuple_contract = tuple()
            if self.args.agent:
                tuple_client += (self.args.agent,)
                tuple_contract += (self.args.agent, self.args.agent)
            if self.args.fond:
                tuple_client += (self.args.fond,)
            if self.args.arhivON:
                tuple_client += (1,)
            if self.args.arhivOFF:
                tuple_client += (0,)
            if self.args.signer:
                tuple_contract += (self.args.signer,)
            if self.args.socium:
                tuple_contract += (None,)
            if self.args.fondStat:
                tuple_contract += (0, 0, 0)
            if self.args.ourStat:
                tuple_contract += (0, 0, 0)
            if self.args.suff:
                tuple_contract += (self.args.suff,)
            tuple_contract += (client_id,)
            tuple_client += (client_id,)
            tuples_clients.append(tuple_client)
            tuples_contracts.append(tuple_contract)
            if i_tek and not (i_tek % 1000):
                if self.leSQLcl:
                    cursor.executemany(self.leSQLcl, tuples_clients)
                    dbconn.commit()
                    tuples_clients = []
                if self.leSQLco:
                    cursor.executemany(self.leSQLco, tuples_contracts)
                    dbconn.commit()
                    tuples_contracts = []
            i_tek += 1
        if self.leSQLcl:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), self.leSQLcl])
            ws_log.append([datetime.now().strftime("%H:%M:%S")] + list(tuples_clients[0]))
            if len(tuples_clients):
                cursor.executemany(self.leSQLcl, tuples_clients)
                dbconn.commit()
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'update clients отработал'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'clients - нет запроса'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'clients - нет данных'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'clients - запрос не исполнен'])
        if self.leSQLco:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), self.leSQLco])
            ws_log.append([datetime.now().strftime("%H:%M:%S")] + list(tuples_contracts[0]))
            if len(tuples_contracts):
                cursor.executemany(self.leSQLco, tuples_contracts)
                dbconn.commit()
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'update contracts отработал'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'contracts - нет запроса'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'contracts - нет данных'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'contracts - запрос не исполнен'])
        wb_log.save(log_name)

if __name__ == '__main__':
    all = my()
    all.one()

