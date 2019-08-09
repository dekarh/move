# -*- coding: utf-8 -*-
# для поиска по базе адресов нужно стартовать сервисы sphinx и fias

from subprocess import Popen, PIPE
import os
import sys
import re
import string
import bz2
from string import digits
from random import random
from dateutil.parser import parse
from collections import OrderedDict

from datetime import datetime, timedelta, time
import time
import pytz
utc=pytz.UTC

import openpyxl
from openpyxl import Workbook
import requests, json


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QDate, QDateTime, QSize, Qt, QByteArray, QTimer, QUrl, QThread
from PyQt5.QtGui import QPixmap, QIcon
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QMainWindow, QWidget, QFrame, QFileDialog, QComboBox

from mysql.connector import MySQLConnection, Error

from move_win import Ui_Form

# import NormalizeFields as norm
from lib import read_config, l, s, fine_phone, format_phone

MANIPULATE_LABELS = ["-------------------------"
                    , "ФИО из поля"
                    , "-------------------------"
                    , "ФИО при рождении из поля"
                    , "-------------------------"
                    , "Регистрация -> Регион"
                    , "Регистрация -> Район"
                    , "Регистрация -> Город"
                    , "Регистрация -> Населенный_пункт"
                    , "Регистрация -> Улица"
                    , "-------------------------"
                    , "Проживание -> Регион"
                    , "Проживание -> Район"
                    , "Проживание -> Город"
                    , "Проживание -> Населенный_пункт"
                    , "Проживание -> Улица"
                    , "-------------------------"
                    , "Адрес регистрации из_поля"
                    , "-------------------------"
                    , "Адрес проживания из поля"
                    , "-------------------------"
                    , "Регион регистрации из номера"
                    , "-------------------------"
                    , "Регион проживания из номера"
                    , "-------------------------"
                    , "Cерия и Номер паспорта из поля"
                    , "-------------------------"
                    , "Генератор некорректных СНИЛС"
                    #, "Пол_получить_из_ФИО"
                    #, "Пол_подставить_свои_значения"
                     ]

SNILS_LABEL = ["СНИЛС"]
FIO_LABELS = ["ФИО.Фамилия", "ФИО.Имя", "ФИО.Отчество"]
FIO_BIRTH_LABELS = ["ФИО_при_рождении.Фамилия", "ФИО_при_рождении.Имя", "ФИО_при_рождении.Отчество"]
FIO_SNILS_LABELS = ["Фамилия_по_СНИЛС", "Имя_по_СНИЛС", "Отчество_по_СНИЛС"]
GENDER_LABEL = ["Пол"]
DATE_BIRTH_LABEL = ["Дата_рождения"]
PLACE_BIRTH_LABELS = ["Место_рождения.Страна", "Место_рождения.Область", "Место_рождения.Район",
                      "Место_рождения.Город"]
PASSPORT_DATA_LABELS = ["Данные_паспорта.Серия", "Данные_паспорта.Номер", "Данные_паспорта.Дата_выдачи",
                        "Данные_паспорта.Кем_выдан", "Данные_паспорта.Код_подразделения"]
ADRESS_REG_LABELS = ["Адрес_регистрации.Индекс",
                     "Адрес_регистрации.Регион", "Адрес_регистрации.Тип_региона",
                     "Адрес_регистрации.Район", "Адрес_регистрации.Тип_района",
                     "Адрес_регистрации.Город", "Адрес_регистрации.Тип_города",
                     "Адрес_регистрации.Населенный_пункт", "Адрес_регистрации.Тип_населенного_пункта",
                     "Адрес_регистрации.Улица", "Адрес_регистрации.Тип_улицы",
                     "Адрес_регистрации.Дом",
                     "Адрес_регистрации.Корпус",
                     "Адрес_регистрации.Квартира"]

ADRESS_LIVE_LABELS = ["Адрес_проживания.Индекс",
                      "Адрес_проживания.Регион", "Адрес_проживания.Тип_региона",
                      "Адрес_проживания.Район", "Адрес_проживания.Тип_района",
                      "Адрес_проживания.Город", "Адрес_проживания.Тип_города",
                      "Адрес_проживания.Населенный_пункт", "Адрес_проживания.Тип_населенного_пункта",
                      "Адрес_проживания.Улица", "Адрес_проживания.Тип_улицы",
                      "Адрес_проживания.Дом",
                      "Адрес_проживания.Корпус",
                      "Адрес_проживания.Квартира"]

PHONES_LABELS = ["Телефон.Мобильный", "Телефон.Родственников", "Телефон.Домашний"]

TECH_LABELS = ["Агент_Ид", "Подписант_Ид", "Пред_Страховщик_Ид"]

FIELDS_IN_RESULT_TABLE_FULL = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, FIO_SNILS_LABELS, GENDER_LABEL,
                                   DATE_BIRTH_LABEL, PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS,
                                   ADRESS_LIVE_LABELS, PHONES_LABELS, TECH_LABELS, MANIPULATE_LABELS]
FIELDS_IN_RESULT_TABLE_SHORT = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, FIO_SNILS_LABELS, GENDER_LABEL,
                                   DATE_BIRTH_LABEL, PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS,
                                   ADRESS_LIVE_LABELS, PHONES_LABELS, TECH_LABELS]

HEAD_RESULT_EXCEL_FILE = ['СНИЛС',
                          'Фамилия', 'Имя', 'Отчество',
                          'Фамилия_при_рождении', 'Имя_при_рождении', 'Отчество_при_рождении',
                          'Фамилия_по_СНИЛС', 'Имя_по_СНИЛС', 'Отчество_по_СНИЛС',
                          'Пол(0_мужской,1_женский)',
                          'Дата_рождения',
                          'Страна_рождения', 'Область_рождения', 'Район_рождения', 'Город_рождения',
                          'Паспорт_серия', 'Паспорт_номер', 'Паспорт_дата', 'Паспорт_Кем выдан',
                          'Паспорт_Код подразделения',

                          'Адрес_регистрации_Индекс',
                          'Адрес_регистрации_Регион', 'Адрес_регистрации_Тип_региона',
                          'Адрес_регистрации_Район', 'Адрес_регистрации_Тип_района',
                          'Адрес_регистрации_Город', 'Адрес_регистрации_Тип_города',
                          'Адрес_регистрации_Населенный_пункт', 'Адрес_регистрации_Тип_населенного_пункта',
                          'Адрес_регистрации_Улица',
                          'Адрес_регистрации_Тип_улицы',
                          'Адрес_регистрации_Дом',
                          'Адрес_регистрации_Корпус',
                          'Адрес_регистрации_Квартира',

                          'Адрес_проживания_Индекс',
                          'Адрес_проживания_Регион', 'Адрес_проживания_Тип_региона',
                          'Адрес_проживания_Район', 'Адрес_проживания_Тип_района',
                          'Адрес_проживания_Город', 'Адрес_проживания_Тип_города',
                          'Адрес_проживания_Населенный_пункт', 'Адрес_проживания_Тип_населенного_пункта',
                          'Адрес_проживания_Улица', 'Адрес_проживания_Тип_улицы',
                          'Адрес_проживания_Дом',
                          'Адрес_проживания_Корпус',
                          'Адрес_проживания_Квартира',

                          'Мобильный_телефон', 'Телефон_родственников', 'Телефон_домашний',
                          'Агент_Ид', 'Подписант_Ид', 'Пред_Страховщик_Ид'
                          ]

LEN_SNILS = 11
LEN_PASSPORT_NOMER = 6
LEN_INDEX_NOMER = 6
LEN_PASSPORT_COD = 6
EAST_GENDER = ['кызы', 'оглы']

########################################################################################################################
# НУЛЕВЫЕ ЗНАЧЕНИЯ
NULL_VALUE = '\\N'  # НУЛЕВОЕ ЗНАЧЕНИЕ В ФАЙЛЕ
NEW_NULL_VALUE = ''  # НОВОЕ НУЛЕВОЕ ЗНАЧЕНИЕ
NEW_NULL_VALUE_FOR_ADDRESS = ['', '', '', '', '', '', '', '', '', '', '', '', '']
NEW_NULL_VALUE_FOR_DATE = '11.11.1111'
NEW_NULL_VALUE_FOR_SERIYA_PASSPORTA = '1111'
NEW_NULL_VALUE_FOR_NOMER_PASSPORTA = '111111'
NEW_NULL_VALUE_FOR_COD_PASSPORTA = '111-111'
NEW_NULL_VALUE_FOR_GENDER = '0'
NEW_NULL_VALUE_FOR_INDEX = '111111'
NEW_NULL_VALUE_FOR_ALL_TEXT = 'заполнить'
NEW_NULL_VALUE_FOR_HOME = 'заполнить'
########################################################################################################################
# ЗНАЧЕНИЕ ПРИ ОШИБКЕ
ERROR_VALUE = 'ERROR'
########################################################################################################################
# СОКРАЩЕНИЯ ТИПОВ В АДРЕСЕ
SPLIT_FIELDS = ['.', ',', ' ', ';', '№', '(', ')']
SPLIT_FIELD = SPLIT_FIELDS[1]
#SPLIT_FIELD = '_x0003_'  # Разделитель для адреса в одной строке (бывает '_x0003_')

# Для варианта когда все поля по раздельности и перепутаны. Используется в FullAdress.get_values. Справочно:
# ['Индекс', 'Регион', 'Тип_региона', 'Район', 'Тип_района', 'Город', 'Тип_города',
#  'Населенный_пункт', 'Тип_населенного_пункта', 'Улица', 'Тип_улицы', 'Дом', 'Корпус', 'Квартира']
ORDER_FIELD = [13, 0, 8, 1, 9, 2, 10, 3, 11, 4, 12, 5, 6, 7]

REG_TYPES = ['обл', 'о', 'область', 'респ', 'республика', 'край', 'кр', 'ар', 'ао', 'авт окр', 'автономный округ',
             'авт обл', 'автономная область', 'город федерального значения', 'гфз']

DISTRICT_TYPES = ['р-н', 'р', 'район']

CITY_TYPES = ['г', 'гор', 'город']

NP_TYPES = ['пгт', 'поселок городского типа',  'посёлок городского типа', 'пос', 'поселение', 'поселок', 'посёлок',
            'п', 'рп', 'рабочий посёлок', 'рабочий поселок', 'кп', 'курортный посёлок', 'курортный поселок', 'к', 'пс',
            'сс', 'смн', 'вл', 'влад', 'владение', 'дп', 'дачный поселок', 'дачный посёлок', 'садовое товарищество',
            'садоводческое некоммерческое товарищество', 'садоводческое товарищество', 'снт', 'нп', 'пст', 'ж/д_ст',
            'ж/д ст', 'железнодорожная станция', 'с', 'село', 'м', 'д', 'дер', 'деревня', 'сл', 'ст', 'ст-ца',
            'станица', 'х', 'хут', 'хутор', 'рзд', 'у', 'урочище', 'клх', 'колхоз', 'свх', 'совхоз', 'зим', 'зимовье',
            'микрорайон', 'мкр']

STREET_TYPES = ['аллея', 'а', 'бульвар', 'б-р', 'бул', 'в/ч', 'военная часть', 'военный городок', 'городок', 'гск',
                'гаражно-строительный кооператив', 'гк', 'гаражный кооператив', 'кв-л', 'квартал', 'линия', 'лин',
                'наб', 'набережная', 'переулок', 'пер', 'переезд', 'пл', 'площадь', 'пр-кт', 'проспект', 'пр',
                'проезд', 'тер', 'терр', 'территория', 'туп', 'тупик', 'ул', 'улица', 'ш', 'шоссе']

HOUSE_CUT_NAME = ['дом', 'д']
CORPUS_CUT_NAME = ['корп', 'корпус', 'стр', 'строение']
APARTMENT_CUT_NAME = ['кв', 'квартира', 'оф', 'офис', 'ап', 'аппартаменты']

ADRESS_TYPES = {
'обл': 1, 'о': 1, 'область': 1, 'респ': 1, 'республика': 1, 'край': 1, 'кр': 1, 'ар': 1, 'ао': 1, 'авт окр': 1, 'автономный округ': 1, 'авт обл': 1, 'автономная область': 1, 'город федерального значения': 1, 'гфз': 1,
'р-н': 3, 'р': 3, 'район': 3,
'г': 5, 'гор': 5, 'город': 5,
'пгт': 7, 'поселок городского типа': 7,  'посёлок городского типа': 7, 'пос': 7, 'поселение': 7, 'поселок': 7, 'посёлок': 7, 'п': 7, 'рп': 7, 'рабочий посёлок': 7, 'рабочий поселок': 7, 'кп': 7, 'курортный посёлок': 7, 'курортный поселок': 7, 'к': 7, 'пс': 7, 'сс': 7, 'смн': 7, 'вл': 7, 'влад': 7, 'владение': 7, 'дп': 7, 'дачный поселок': 7, 'дачный посёлок': 7, 'садовое товарищество': 7, 'садоводческое некоммерческое товарищество': 7, 'садоводческое товарищество': 7, 'снт': 7, 'нп': 7, 'пст': 7, 'ж/д_ст': 7, 'ж/д ст': 7, 'железнодорожная станция': 7, 'с': 7, 'село': 7, 'м': 7, 'дер': 7, 'деревня': 7, 'сл': 7, 'ст': 7, 'ст-ца': 7, 'станица': 7, 'х': 7, 'хут': 7, 'хутор': 7, 'рзд': 7, 'у': 7, 'урочище': 7, 'клх': 7, 'колхоз': 7, 'свх': 7, 'совхоз': 7, 'зим': 7, 'зимовье': 7, 'микрорайон': 7, 'мкр' : 7,
'аллея': 9, 'а': 9, 'алл': 9, 'бульвар': 9, 'б-р': 9, 'бул': 9, 'блв': 9, 'в/ч': 9, 'военная часть': 9, 'военный городок': 9, 'городок': 9, 'гск': 9, 'гаражно-строительный кооператив': 9, 'гк': 9, 'гаражный кооператив': 9, 'кв-л': 9, 'квартал': 9, 'линия': 9, 'лин': 9, 'наб': 9, 'набережная': 9, 'переулок': 9, 'пер': 9, 'переезд': 9, 'пл': 9, 'площадь': 9, 'пр-кт': 9, 'проспект': 9, 'пр': 9, 'проезд': 9, 'тер': 9, 'терр': 9, 'территория': 9, 'туп': 9, 'тупик': 9, 'ул': 9, 'улица': 9, 'ш': 9, 'шоссе': 9,
'дом': 11,  'д': 11,
'корп': 12,  'корпус': 12,  'стр': 12,  'строение': 12,
'кв': 13,  'квартира': 13,  'оф': 13,  'офис': 13,  'ап': 13,  'аппартаменты': 13
}

ALL_CUT_NAMES = [
    'а обл','а окр','АО','Аобл','г','г','г ф з','гфз','край','обл','обл','округ','Респ','респ','АО','АО','вн тер г',
    'г о', 'го','м р-н','п','пос','р-н','тер','у','у','волость','г','г','дп','кп','массив','п','п/о','пгт','пгт','рп',
    'с/а','с/мо','с/о','с/п','с/с','тер','р-н','тер','аал','автодорога','арбан','аул','волость','высел','г','г-к','гп',
    'д','дп','ж/д б-ка','ж/д пл-ка','ж/д пл-ма','ж/д_будка','ж/д_казарм','ж/д_оп','ж/д_платф','ж/д_пост','ж/д_рзд',
    'ж/д_ст','жилзона','жилрайон','заимка','казарма','кв-л','кордон','кп','лпх','м','массив','мкр','нп','остров','п',
    'п/о','п/р','п/ст','пгт','пгт','погост','починок','промзона','рзд','рп','с','сл','снт','ст','ст-ца','тер','у','х',
    'а/я','ал','аллея','балка','берег','б-р','бугор','вал','взв','въезд','г-к','гск','д','днп','дор','ж/д_будка',
    'ж/д_казарм','ж/д_оп','ж/д_платф','ж/д_пост','ж/д_рзд','ж/д_ст','жт','заезд','ззд','зона','казарма','кв-л','км',
    'кольцо','коса','к-цо','линия','лн','м','маяк','мгстр','местность','мкр','мост','н/п','наб','наб','нп','остров',
    'п','п/о','п/р','п/ст','парк','пер','пер','пер-д','переезд','пл','платф','пл-ка','полустанок','пр-д','пр-к','пр-ка'
    ,'пр-кт','пр-лок','проезд','промзона','просек','просека','проселок','проул','проулок','рзд','рзд','ряд','ряды','с',
    'с/т','сад','сзд','с-к','сквер','сл','снт','спуск','с-р','ст','стр','тер','тер ДНТ','тер СНТ','тракт','туп','ул',
    'ул','уч-к','ф/х','ферма','х','ш','ш','влд','д','двлд','ДОМ','зд','к','кот','ОНС','пав','соор','стр','шахта','г-ж',
    'кв','ком','офис','п-б','подв','помещ','раб уч','скл','торг зал','цех','вн р-н','г п','с п','с/с','а/я','аал','ал',
    'арбан','аул','б-г','б-р','вал','взд','г-к','гск','д','днп','дор','ж/д б-ка','ж/д к-ма','ж/д пл-ма','ж/д рзд',
    'ж/д ст','ж/р','ззд','зона','кв-л','км','коса','к-цо','лн','местность','месторожд','м-ко','мкр','мкр','н/п','наб',
    'ост-в','п','п/р','парк','пер','пер-д','п-к','пл','платф','пл-ка','порт','пр-д','пр-к','пр-ка','пр-кт','пр-лок',
    'промзона','проул','рзд','р-н','с','сад','с-к','сквер','сл','снт','с-р','ст','стр','тер','тер','тер ГСК','тер ДНО',
    'тер ДНП','тер ДНТ','тер ДПК','тер ОНО','тер ОНП','тер ОНТ','тер ОПК','тер СНО','тер СНП','тер СНТ','тер СПК',
    'тер ТСН','тер СОСН','тер ф х','тракт','туп','ул','ус','ф/х','х','ш','ю','з/у','гск','днп','местность','мкр','н/п',
    'промзона','сад','снт','тер','ф/х','а/я','аал','аллея','арбан','аул','берег','б-р','вал','въезд','высел','г-к',
    'гск','д','дор','ж/д_будка','ж/д_казарм','ж/д_оп','ж/д_платф','ж/д_пост','ж/д_рзд','ж/д_ст','жт','заезд','зона',
    'казарма','кв-л','км','кольцо','коса','линия','м','мкр','мост','наб','нп','остров','п','п/о','п/р','п/ст','парк',
    'пер','переезд','пл','платф','пл-ка','починок','пр-кт','проезд','просек','просека','проселок','проулок','рзд',
    'ряды','с','сад','сквер','сл','снт','спуск','ст','стр','тер','тракт','туп','ул','уч-к','ферма','х','ш'
]

q1 = """
'обл': 'REG_TYPES', 'о': 'REG_TYPES', 'область': 'REG_TYPES', 'респ': 'REG_TYPES', 'республика': 'REG_TYPES', 'край': 'REG_TYPES', 'кр': 'REG_TYPES', 'ар': 'REG_TYPES', 'ао': 'REG_TYPES', 'авт окр': 'REG_TYPES', 'автономный округ': 'REG_TYPES', 'авт обл': 'REG_TYPES', 'автономная область': 'REG_TYPES', 'город федерального значения': 'REG_TYPES', 'гфз': 'REG_TYPES',
'р-н': 'DISTRICT_TYPES', 'р': 'DISTRICT_TYPES', 'район': 'DISTRICT_TYPES',
'г': 'CITY_TYPES', 'гор': 'CITY_TYPES', 'город': 'CITY_TYPES',
'пгт': 'NP_TYPES', 'поселок городского типа': 'NP_TYPES',  'посёлок городского типа': 'NP_TYPES', 'пос': 'NP_TYPES', 'поселение': 'NP_TYPES', 'поселок': 'NP_TYPES', 'посёлок': 'NP_TYPES', 'п': 'NP_TYPES', 'рп': 'NP_TYPES', 'рабочий посёлок': 'NP_TYPES', 'рабочий поселок': 'NP_TYPES', 'кп': 'NP_TYPES', 'курортный посёлок': 'NP_TYPES', 'курортный поселок': 'NP_TYPES', 'к': 'NP_TYPES', 'пс': 'NP_TYPES', 'сс': 'NP_TYPES', 'смн': 'NP_TYPES', 'вл': 'NP_TYPES', 'влад': 'NP_TYPES', 'владение': 'NP_TYPES', 'дп': 'NP_TYPES', 'дачный поселок': 'NP_TYPES', 'дачный посёлок': 'NP_TYPES', 'садовое товарищество': 'NP_TYPES', 'садоводческое некоммерческое товарищество': 'NP_TYPES', 'садоводческое товарищество': 'NP_TYPES', 'снт': 'NP_TYPES', 'нп': 'NP_TYPES', 'пст': 'NP_TYPES', 'ж/д_ст': 'NP_TYPES', 'ж/д ст': 'NP_TYPES', 'железнодорожная станция': 'NP_TYPES', 'с': 'NP_TYPES', 'село': 'NP_TYPES', 'м': 'NP_TYPES', 'дер': 'NP_TYPES', 'деревня': 'NP_TYPES', 'сл': 'NP_TYPES', 'ст': 'NP_TYPES', 'ст-ца': 'NP_TYPES', 'станица': 'NP_TYPES', 'х': 'NP_TYPES', 'хут': 'NP_TYPES', 'хутор': 'NP_TYPES', 'рзд': 'NP_TYPES', 'у': 'NP_TYPES', 'урочище': 'NP_TYPES', 'клх': 'NP_TYPES', 'колхоз': 'NP_TYPES', 'свх': 'NP_TYPES', 'совхоз': 'NP_TYPES', 'зим': 'NP_TYPES', 'зимовье': 'NP_TYPES', 'микрорайон': 'NP_TYPES', 'мкр' : 'NP_TYPES',
'аллея': 'STREET_TYPES', 'а': 'STREET_TYPES', 'алл': 'STREET_TYPES', 'бульвар': 'STREET_TYPES', 'б-р': 'STREET_TYPES', 'бул': 'STREET_TYPES', 'блв': 'STREET_TYPES', 'в/ч': 'STREET_TYPES', 'военная часть': 'STREET_TYPES', 'военный городок': 'STREET_TYPES', 'городок': 'STREET_TYPES', 'гск': 'STREET_TYPES', 'гаражно-строительный кооператив': 'STREET_TYPES', 'гк': 'STREET_TYPES', 'гаражный кооператив': 'STREET_TYPES', 'кв-л': 'STREET_TYPES', 'квартал': 'STREET_TYPES', 'линия': 'STREET_TYPES', 'лин': 'STREET_TYPES', 'наб': 'STREET_TYPES', 'набережная': 'STREET_TYPES', 'переулок': 'STREET_TYPES', 'пер': 'STREET_TYPES', 'переезд': 'STREET_TYPES', 'пл': 'STREET_TYPES', 'площадь': 'STREET_TYPES', 'пр-кт': 'STREET_TYPES', 'проспект': 'STREET_TYPES', 'пр': 'STREET_TYPES', 'проезд': 'STREET_TYPES', 'тер': 'STREET_TYPES', 'терр': 'STREET_TYPES', 'территория': 'STREET_TYPES', 'туп': 'STREET_TYPES', 'тупик': 'STREET_TYPES', 'ул': 'STREET_TYPES', 'улица': 'STREET_TYPES', 'ш': 'STREET_TYPES', 'шоссе': 'STREET_TYPES',
'дом': 'HOUSE_CUT_NAME',  'д': 'HOUSE_CUT_NAME',
'корп': 'CORPUS_CUT_NAME',  'корпус': 'CORPUS_CUT_NAME',  'стр': 'CORPUS_CUT_NAME',  'строение': 'CORPUS_CUT_NAME',
'кв': 'APARTMENT_CUT_NAME',  'квартира': 'APARTMENT_CUT_NAME',  'оф': 'APARTMENT_CUT_NAME',  'офис': 'APARTMENT_CUT_NAME',  'ап': 'APARTMENT_CUT_NAME',  'аппартаменты': 'APARTMENT_CUT_NAME'
"""
########################################################################################################################
# ЗНАЧЕНИЕ В ПОЛЕ "ПОЛ" ИЗМЕНЯЕМ В ПРОЦЕССЕ
#female_gender_value = 'Ж'
#male_gender_value = 'М'
#gender_length = 1
########################################################################################################################
# ЗАПОЛНЕНИЕ Агент_Ид, Подписант_Ид, Пред_Страховщик_Ид
#AGENT_ID = '10061'
#AGENT_ID = '9954'
#AGENT_ID = '9986'
#PODPISANT_ID = '208'
PREDSTRAH_ID = '1'
########################################################################################################################
# ИМЕНА ДЛЯ КЛЮЧЕЙ СЛОВАРЕЙ И ДЛЯ ПОРЯДКА ВЫВОД СЛОВАРЯ

FULL_ADRESS_LABELS = ['Индекс', 'Регион', 'Тип_региона', 'Район', 'Тип_района', 'Город', 'Тип_города',
                      'Населенный_пункт', 'Тип_населенного_пункта', 'Улица', 'Тип_улицы', 'Дом', 'Корпус', 'Квартира']

PASSPORT_LABELS = ['Серия', 'Номер', 'Дата_выдачи', 'Кем_выдан', 'Код_подразделения']

FIO_KEY_LABELS = ['Фамилия', 'Имя', 'Отчество']

BIRTH_PLACE_LABELS = ['Страна', 'Область', 'Район', 'Город']
########################################################################################################################
REGIONS = [
    "", "Адыгея респ.", "Башкортостан респ.", "Бурятия респ.", "Алтай респ.", "Дагестан респ.", "Ингушетия респ.",
    "Кабардино-Балкарская респ.", "Калмыкия респ.", "Карачаево-Черкесская респ.", "Карелия респ.", "Коми респ.",
    "Марий_Эл респ.", "Мордовия респ.", "Саха/Якутия/ респ.", "Северная_Осетия-Алания респ.", "Татарстан респ.",
    "Тыва респ.", "Удмуртская респ.", "Хакасия респ.", "Чеченская респ.", "Чувашская респ.", "Алтайский край",
    "Краснодарский край", "Красноярский край", "Приморский край", "Ставропольский край", "Хабаровский край",
    "Амурская обл.", "Архангельская обл.", "Астраханская обл.", "Белгородская обл.", "Брянская обл.",
    "Владимирская обл.", "Волгоградская обл.", "Вологодская обл.", "Воронежская обл.", "Ивановская обл.",
    "Иркутская обл.", "Калининградская обл.", "Калужская обл.", "Камчатский край", "Кемеровская обл.", "Кировская обл.",
    "Костромская обл.", "Курганская обл.", "Курская обл.", "Ленинградская обл.", "Липецкая обл.", "Магаданская обл.",
    "Московская обл.", "Мурманская обл.", "Нижегородская обл.", "Новгородская обл.", "Новосибирская обл.",
    "Омская обл.", "Оренбургская обл.", "Орловская обл.", "Пензенская обл.", "Пермский край", "Псковская обл.",
    "Ростовская обл.", "Рязанская обл.", "Самарская обл.", "Саратовская обл.", "Сахалинская обл.", "Свердловская обл.",
    "Смоленская обл.", "Тамбовская обл.", "Тверская обл.", "Томская обл.", "Тульская обл.", "Тюменская обл.",
    "Ульяновская обл.", "Челябинская обл.", "Забайкальский край", "Ярославская обл.", "Москва г.", "Санкт-Петербург г.",
    "Еврейская авт.обл.", "Агинский_Бурятский авт.округ", "Коми-Пермяцкий авт.округ", "Корякский авт.округ",
    "Ненецкий авт.округ", "Таймырский_(Долгано-Ненецкий) авт.округ", "Усть-Ордынский_Бурятский авт.округ",
    "Ханты-Мансийский/Югра/ авт.округ", "Чукотский авт.округ", "Эвенкийский авт.округ", "Ямало-Ненецкий авт.округ",
    "","Крым респ.", "Севастополь г.","","","","","","","Байконур г." ]
########################################################################################################################
# True - заменять СНИЛС на некорректный, неиспользованный в Сатурн

GENERATE_SNILS = False

########################################################################################################################

IN_IDS = ['ID','ИД_КЛИЕНТА','CLIENT_ID']
IN_SNILS = ['СНИЛС', 'СТРАХОВОЙ_НОМЕР', 'СТРАХОВОЙ НОМЕР', 'NUMBER','СТРАХОВОЙНОМЕР']
IN_NAMES = ['ID', 'СНИЛС', 'СТРАХОВОЙ_НОМЕР', 'СТРАХОВОЙ НОМЕР', 'NUMBER', 'ФАМИЛИЯ', 'ИМЯ', 'ОТЧЕСТВО', 'ФИО']

DIR4MOVE = '/home/da3/Move/'
DIR4IMPORT = '/home/da3/CheckLoad/'
DIR4CFGIMPORT = '/home/da3/CheckLoad/cfg/'
DIR4PCHECK = '/home/da3/PasportChecks/'
DIR4DELDOUBLESPHONES = '/home/da3/DelDoublesPhones/'


class MainWindowSlots(Ui_Form):   # Определяем функции, которые будем вызывать в слотах

    def setupUi(self, form):
        Ui_Form.setupUi(self,form)
        self.partner = 0
        self.phones = []
        self.passports = {}
        self.tableWidget.setColumnCount(2)
        self.tableWidget.setHorizontalHeaderLabels(('Результат', 'Исходник'))
        for j in range(self.tableWidget.columnCount()):
            self.tableWidget.setColumnWidth(j, 220)
        self.tableWidget.setRowCount(0)
        self.dbconfig = read_config(filename='move.ini', section='mysql')
        self.dbconfig_pasp = read_config(filename='move.ini', section='pasport')
        self.MoveImportPasport = 1
        self.signer_ids = []
        self.signer_names = {}
        self.fond_ids = []
        self.fond_names = {}
        self.agent_ids = []
        self.agent_names = {}
        self.clients_ids = []
        self.fond_touched = False
        self.fonds_str = '1'
        self.agent_touched = False
        self.signer_touched = False
        self.cfg_file_touched = False
        self.cfg_file_loaded = False
        self.cfg_file_names = {}
        self.file_touched = False
        self.file_loaded = False
        self.table_loaded = False
        self.file_names = {}
        self.file_name = ''
        self.tab_names = {}
        self.table = []
        self.twParsingResult.hide()
        self.cmbGenderType.addItems(['М или Ж', '0 или 1', 'Мужской или Женский'])
        self.cmbParsingType.addItems(['стандартный', 'перемешаный', 'КЛАДР'])
        self.refresh()
        dbconfig = read_config(filename='move.ini', section='mysql')
        dbconn = MySQLConnection(**dbconfig)
        dbcursor = dbconn.cursor()
        dbcursor.execute('SELECT min(`number`) FROM  clients WHERE `number` > 99900000000;')
        dbrows = dbcursor.fetchall()
        dbconn.close()
        self.start_snils = int('{0:011d}'.format(dbrows[0][0])[:-2])    # 9 цифр неправильного СНИЛСа с которого уменьшаем
        self.start_snils_cs = int('{0:011d}'.format(dbrows[0][0])[-2:]) # контрольная сумма неправильного СНИЛСа
        self.has_gen_snils = False
        return

    def checksum(self, snils_dig):  # Вычисляем 2 последних цифры СНИЛС по первым 9-ти
        def snils_csum(sn):
            k = range(9, 0, -1)
            pairs = zip(k, [int(x) for x in sn.replace('-', '').replace(' ', '')])
            return sum([k * v for k, v in pairs])
        snils = '{0:09d}'.format(snils_dig)
        csum = snils_csum(snils)
        while csum > 101:
            csum %= 101
        if csum in (100, 101):
            csum = 0
        return csum

    def refresh(self):
        if self.fond_touched:                                   # Запомнили позиции combobox'ов
            fond_id = self.fond_ids[self.cmbFond.currentIndex()]
        if self.agent_touched:
            agent_id = self.agent_ids[self.cmbAgent.currentIndex()]
        if self.signer_touched:
            signer_id = self.signer_ids[self.cmbSigner.currentIndex()]
        if self.cfg_file_loaded:
            cfg_file_name = self.cmbCfgFile.currentText()
        if self.file_loaded:
            file_name = self.cmbFile.currentText()
            tab_name = self.cmbTab.currentText()
        self.selectAction()

        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        if self.leAgent.text().strip():
            sql = "SELECT CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, '-', position_id), code " \
                  "FROM offices_staff " \
                  "WHERE CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, user_lastname) LIKE %s " #\
                  #"AND user_fired = 0"
            cursor.execute(sql, ('%' + self.leAgent.text() + '%',))
        else:
            sql = "SELECT CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, '-', position_id), code " \
                  "FROM offices_staff " #\
                  #"WHERE user_fired = 0"
            cursor.execute(sql)
        rows = cursor.fetchall()
        agents = []
        self.agent_names = {}
        self.agent_ids = []
        for i, row in enumerate(rows):
            agents.append(row[0])
            self.agent_names[row[1]] = row[0]
            self.agent_ids.append(row[1])
        self.cmbAgent.clear()
        self.cmbAgent.addItems(agents)
        cursor = dbconn.cursor()
        if self.leFond.text().strip():
            sql = "SELECT CONCAT_WS(' ', id, '-', name), id FROM subdomains " \
                  "WHERE CONCAT_WS(' ', id, '-', name) LIKE %s AND id IN (2,6,8,11,12,13,14)"
            cursor.execute(sql, ('%' + self.leFond.text() + '%',))
        else:
            sql = "SELECT CONCAT_WS(' ', id, '-', name), id FROM subdomains WHERE id IN (2,6,8,11,12,13,14)"
            cursor.execute(sql)
        rows = cursor.fetchall()
        fonds = []
        self.fond_ids = []
        self.fond_names = {}
        self.fonds_str = '1'
        for i, row in enumerate(rows):
            fonds.append(row[0])
            self.fond_names[row[1]] = row[0]
            self.fond_ids.append(row[1])
            if i == 0:
                self.fonds_str = str(row[1])
                continue
            self.fonds_str += ','+ str(row[1])
        self.cmbFond.clear()
        self.cmbFond.addItems(fonds)

        cursor = dbconn.cursor()
        if self.leSigner.text().strip():
            if self.fond_touched:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers " \
                      "WHERE CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname) LIKE %s " \
                      "AND subdomain_id = %s"
                cursor.execute(sql, (self.fond_ids[self.cmbFond.currentIndex()],))
            else:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers " \
                      "WHERE CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname) LIKE %s " \
                      "AND subdomain_id IN (2,6,8,11,12,13,14) AND subdomain_id IN (" + self.fonds_str + ")"
                cursor.execute(sql, ('%' + self.leSigner.text() + '%',))
        else:
            if self.fond_touched:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers WHERE subdomain_id = %s"
                cursor.execute(sql, (self.fond_ids[self.cmbFond.currentIndex()],))
            else:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers WHERE subdomain_id IN (" + self.fonds_str + ") " \
                      "AND subdomain_id IN (2,6,8,11,12,13,14)"
                cursor.execute(sql)
        rows = cursor.fetchall()
        self.signer_ids = []
        self.signer_names = {}
        signers = []
        for i, row in enumerate(rows):
            self.signer_names[row[1]] = row[0]
            self.signer_ids.append(row[1])
            signers.append(row[0])
        self.cmbSigner.clear()
        self.cmbSigner.addItems(signers)

        rows = sorted(os.listdir(DIR4CFGIMPORT))                        # список конфигов
        cfg_files = []
        for row in rows:
            if row.find('.xlsx') > -1:
                cfg_files.append(row)
        self.cfg_file_names = {}
        for i, cfg_file in enumerate(cfg_files):
            self.cfg_file_names[cfg_file] = i
        self.cmbCfgFile.clear()
        self.cmbCfgFile.addItems(cfg_files)

        try:                                                    # список файлов загрузки
            files = sorted(os.listdir(path=self.leDir.text()))
            for i, file in enumerate(files):
                self.file_names[file] = i
            self.cmbFile.clear()
            self.cmbFile.addItems(files)
            try:
                if not self.file_loaded:
                    self.twAllExcels.setColumnCount(0)
                    self.twAllExcels.setRowCount(0)
                    self.file_name = ''
                else:
                    self.cmbFile.setCurrentIndex(self.file_names[file_name])
            except ValueError:
                self.file_loaded = False
                self.twAllExcels.setColumnCount(0)
                self.twAllExcels.setRowCount(0)
                self.file_name = ''
            else:
                if self.file_loaded:
                    if self.cmbFile.currentText()[len(self.cmbFile.currentText()) - 5:] == '.xlsx':
                        self.wb = openpyxl.load_workbook(filename=self.leDir.text() + self.cmbFile.currentText(),
                                                         read_only=True)
                        tabs = self.wb.sheetnames
                        for i, tab in enumerate(tabs):
                            self.tab_names[tab] = i
                        self.cmbTab.clear()
                        self.cmbTab.addItems(tabs)
        except OSError:
            self.errMessage('Нет такого файла')

        try:                                            # Восстанавиваем позиции combobox'ов
            if self.fond_touched:
                self.cmbFond.setCurrentIndex(self.fond_ids.index(fond_id))
        except ValueError:
            self.fond_touched = False
        try:
            if self.agent_touched:
                self.cmbAgent.setCurrentIndex(self.agent_ids.index(agent_id))
        except ValueError:
            self.agent_touched = False
        try:
            if self.signer_touched:
                self.cmbSigner.setCurrentIndex(self.signer_ids.index(signer_id))
        except ValueError:
            self.signer_touched = False
        try:
            if self.file_loaded:
                self.cmbTab.setCurrentIndex(self.tab_names[tab_name])
        except ValueError:
            self.file_loaded = False
            self.twAllExcels.setColumnCount(0)
            self.twAllExcels.setRowCount(0)
            self.file_name = ''
        try:
            if self.cfg_file_loaded:
                self.cmbCfgFile.setCurrentIndex(self.cfg_file_names[cfg_file_name])
        except ValueError:
            self.cfg_file_loaded = False
        return

    def click_pbRefresh(self):
        self.refresh()
        return

    def click_pbMove(self):
        if not self.file_touched:                               # Проверяем достаточность данных
            self.frFile.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leAgent.isEnabled() and not self.agent_touched:
            self.frAgent.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leFond.isEnabled() and not self.fond_touched:
            self.frFond.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leSigner.isEnabled() and not self.signer_touched:
            self.frSigner.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return

        if self.chbServerString.isChecked():
            cmd = 'python3 move_cmd.py ' + (self.leDir.text() + self.cmbFile.currentText()).replace(' ','\ ') \
                  +' -sheetName ' + self.cmbTab.currentText().replace(' ', '\ ')
            if self.leFond.isEnabled():
                cmd += ' -fond ' + str(self.fond_ids[self.cmbFond.currentIndex()])
            if self.leAgent.isEnabled():
                cmd += ' -agent ' + str(self.agent_ids[self.cmbAgent.currentIndex()])
            if self.leSigner.isEnabled():
                cmd += ' -signer ' + str(self.signer_ids[self.cmbSigner.currentIndex()])
            if self.chbClientOnly.isChecked():
                cmd += ' -clientOnly True'
            if self.chbSocium.isChecked():
                cmd += ' -socium True'
            if self.chbSuff.isChecked():
                cmd += ' -suff ' + self.leSuff.text()
            if self.chbOurStat.isChecked():
                cmd += ' -ourStat True'
            if self.chbFondStat.isChecked():
                cmd += ' -fondStat True'
            if self.chbArhivON.isChecked():
                cmd += ' -arhivON True'
            if self.chbArhivOFF.isChecked():
                cmd += ' -arhivOFF True'
            if self.chbNoDubPhonePartner.isChecked():
                cmd += ' -noDubPhonePartner True'
            if self.chbNoBackup.isChecked():
                cmd += ' -noBackup True'
            self.leSQLcl.setText(cmd)
            self.leSQLco.setText(cmd)
            return

        # Создаем файл с исходными данными и логом
        wb_log = openpyxl.Workbook(write_only=True)

        ws_log = wb_log.create_sheet('Лог')
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Начинаем'])

        log_name = DIR4MOVE + datetime.now().strftime('%Y-%m-%d_%H-%M')
        if self.fond_touched:
            log_name += 'ф' + str(self.fond_ids[self.cmbFond.currentIndex()])
        if self.agent_touched:
            log_name += 'а' + str(self.agent_ids[self.cmbAgent.currentIndex()])
        log_name += '.xlsx'

        if not self.chbNoBackup.isChecked():
            all_clients_ids = "'" + self.clients_ids[0] + "'"       # Проверка на дубли clients
            for i, client_id in enumerate(self.clients_ids):
                if i == 0:
                    continue
                all_clients_ids += ",'" + client_id + "'"
            sql = "SELECT cl.client_id FROM clients AS cl WHERE cl.client_id IN (" + all_clients_ids + \
                  ") GROUP BY cl.client_id HAVING COUNT(cl.client_id) > 1 ORDER BY cl.client_id DESC"
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
            exit_because_doubles = False
            if len(rows) > 0:
                exit_because_doubles = True
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в clients'])
                ws_clients = wb_log.create_sheet('Дубли в clients')
                for row in rows:
                    ws_clients.append([row[0]])
            else:
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' В clients нет дублей'])
                                                                    # Проверка на дубли contracts
            sql = "SELECT co.client_id FROM contracts AS co WHERE co.client_id IN (" + all_clients_ids + \
                  ") GROUP BY co.client_id HAVING COUNT(co.client_id) > 1 ORDER BY co.client_id DESC"
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
            if len(rows) > 0:
                exit_because_doubles = True
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в contracts'])
                ws_contracts = wb_log.create_sheet('Дубли в contracts')
                for row in rows:
                    ws_contracts.append([row[0]])
            else:
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' В contracts нет дублей'])
            if exit_because_doubles:  # Если дубли в clients или contracts - ничего не переносим
                ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Аварийное завершение - дублирование записей'])
                return

        # Проверка на дубли исходной таблицы
        doubles_in_input = list(set([x for x in self.clients_ids if self.clients_ids.count(x) > 1]))
        if len(doubles_in_input) > 0:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в исходной таблице'])
            ws_input_doubles = wb_log.create_sheet('Дубли в исходной таблице')
            for row in doubles_in_input:
                ws_input_doubles.append(row)
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), ' В исходной таблице нет дублей'])

        if not self.chbNoBackup.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Дублируем исходную excel таблицу в этот файл'])
            ws_input = wb_log.create_sheet('Исходная таблица')
            for table_row in self.table:
                row = []
                for cell in table_row:
                    row.append(cell)
                ws_input.append(row)
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Бэкап исходного состояния БД создан'])
            ws_backup = wb_log.create_sheet('бэкап БД')
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            sql = "SELECT cl.*, co.* FROM clients AS cl LEFT JOIN contracts AS co " \
                  "ON (cl.client_id = co.client_id) WHERE cl.client_id IN (" + all_clients_ids + ")"
            cursor.execute(sql)
            dbrows = cursor.fetchall()
            ws_backup.append(cursor.column_names)
            for dbrow in dbrows:
                row = []
                for dbcell in dbrow:
                    row.append(dbcell)
                ws_backup.append(row)

        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Состояние программы:'])
        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'файл ', self.file_name])
        if self.leFond.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', self.cmbFond.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', 'не выбран'])
        if self.leAgent.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', self.cmbAgent.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', 'не выбран'])
        if self.leSigner.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', self.cmbSigner.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', 'не выбран'])
        if self.chbClientOnly.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Перенести только клиента', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Перенести только клиента', 'не выбрано'])
        if self.chbSocium.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить номер Социума', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить номер Социума', 'не выбрано'])
        if self.chbSuff.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Суффикс', self.leSuff.text()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Суффикс', 'не выбрано'])
        if self.chbOurStat.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить внутренние статусы', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить внутренние статусы', 'не выбрано'])
        if self.chbFondStat.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить статусы Фонда', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Сбросить статусы Фонда', 'не выбрано'])
        if self.chbArhivON.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Поставить флаг "Архивный"', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Поставить флаг "Архивный"', 'не выбрано'])
        if self.chbArhivOFF.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Убрать флаг "Архивный"', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Убрать флаг "Архивный"', 'не выбрано'])
        if self.chbNoDubPhonePartner.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Без дублей телефонов у партнера', 'выбрано'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Без дублей телефонов у партнера', 'не выбрано'])

        # Список телефонов у партнера в фонде в который переносим
        if self.chbNoDubPhonePartner.isChecked() and self.leAgent.isEnabled():
            dbconn = MySQLConnection(**self.dbconfig)
            cursor = dbconn.cursor()
            cursor.execute('SELECT partner_code FROM offices_staff WHERE code = %s',
                           (self.agent_ids[self.cmbAgent.currentIndex()],))
            partner = cursor.fetchall()
            if self.partner != partner[0][0]:
                self.partner = partner[0][0]
                phones = []
                cursor = dbconn.cursor()
                sql_tel = 'SELECT phone_personal_mobile FROM clients AS cl LEFT JOIN offices_staff AS os ' \
                          'ON cl.inserted_user_code = os.code WHERE os.partner_code = %s'
                if self.leFond.isEnabled():
                    cursor.execute(sql_tel +' AND cl.subdomain_id = %s', (partner[0][0],
                                                                          self.fond_ids[self.cmbFond.currentIndex()]))
                else:
                    cursor.execute(sql_tel, (partner[0][0],))
                phones_sql = cursor.fetchall()
                self.progressBar.setMaximum(len(phones_sql) - 1)
                for i, phone_sql in enumerate(phones_sql):
                    if not (i % 10000):
                        self.progressBar.setValue(i)
                    if phone_sql[0] and phone_sql[0] not in phones:
                        phones.append(phone_sql[0])
                    #if i > 10000:
                    #    break
                self.phones = phones
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Формируем запросы:'])
        sql_cl = 'UPDATE clients AS cl SET'
        sql_co = 'UPDATE contracts AS co SET'
        if self.leAgent.isEnabled():
            sql_cl += ' cl.inserted_user_code = %s'
            sql_co += ' co.inserted_code = %s, co.agent_code = %s'
        if self.leFond.isEnabled():
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ','
            sql_cl += ' cl.subdomain_id = %s'
        if self.chbArhivON.isChecked() or self.chbArhivOFF.isChecked():
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ','
            sql_cl += ' cl.archived = %s'
        if self.leSigner.isEnabled():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.signer_id = %s'
        if self.chbSocium.isChecked():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.socium_contract_number = %s'
        if self.chbFondStat.isChecked():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.external_status_code = %s, co.external_status_secure_code = %s,' \
                      ' co.external_status_callcenter_code = %s'
        if self.chbOurStat.isChecked():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.status_code = %s, co.status_secure_code = %s, co.status_callcenter_code = %s'
        if self.chbSuff.isChecked():
            if sql_co[len(sql_co) - 2:] == '%s':
                sql_co += ','
            sql_co += ' co.partner_remote_id = %s'
        if sql_cl[len(sql_cl) - 3:] == 'SET':
            self.leSQLcl.setText('')
        else:
            self.leSQLcl.setText(sql_cl + ' WHERE cl.client_id = %s')
        if sql_co[len(sql_co) - 3:] == 'SET' or self.chbClientOnly.isChecked():
            self.leSQLco.setText('')
        else:
            self.leSQLco.setText(sql_co + ' WHERE co.client_id = %s')

        tuples_clients = []                                     # Формируем переменные для запросов
        tuples_contracts = []
        self.progressBar.setMaximum(len(self.clients_ids)-1)
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        cursor_phones = dbconn.cursor()
        i_tek = 0
        if len(self.phones):
            ws_phones_doubles = wb_log.create_sheet('Дубли телефонов у партнера')
            ws_phones_doubles.append(['client_id', 'СНИЛС', 'телефон'])
        for i, client_id in enumerate(self.clients_ids):
            if len(self.phones):
                cursor_phones.execute('SELECT phone_personal_mobile, number FROM clients AS cl WHERE cl.client_id = %s',
                                      (client_id,))
                rows = cursor_phones.fetchall()
                if rows[0][0] in self.phones:
                    ws_phones_doubles.append([client_id, rows[0][1], rows[0][0]])
                    continue
            tuple_client = tuple()
            tuple_contract = tuple()
            if self.leAgent.isEnabled():
                tuple_client += (self.agent_ids[self.cmbAgent.currentIndex()],)
                tuple_contract += (self.agent_ids[self.cmbAgent.currentIndex()],
                                   self.agent_ids[self.cmbAgent.currentIndex()])
            if self.leFond.isEnabled():
                tuple_client += (self.fond_ids[self.cmbFond.currentIndex()],)
            if self.chbArhivON.isChecked():
                tuple_client += (1,)
            if self.chbArhivOFF.isChecked():
                tuple_client += (0,)
            if self.leSigner.isEnabled():
                tuple_contract += (self.signer_ids[self.cmbSigner.currentIndex()],)
            if self.chbSocium.isChecked():
                tuple_contract += (None,)
            if self.chbFondStat.isChecked():
                tuple_contract += (0,0,0)
            if self.chbOurStat.isChecked():
                tuple_contract += (0,0,0)
            if self.chbSuff.isChecked():
                tuple_contract += (self.leSuff.text(),)
            tuple_contract += (client_id,)
            tuple_client += (client_id,)
            tuples_clients.append(tuple_client)
            tuples_contracts.append(tuple_contract)
            if i_tek and not (i_tek % 1000):
                self.progressBar.setValue(i)
                if self.leSQLcl.text():
                    cursor.executemany(self.leSQLcl.text(), tuples_clients)
                    dbconn.commit()
                    tuples_clients = []
                if self.leSQLco.text():
                    cursor.executemany(self.leSQLco.text(), tuples_contracts)
                    dbconn.commit()
                    tuples_contracts = []
            i_tek += 1
        if self.leSQLcl.text():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), self.leSQLcl.text()])
            ws_log.append([datetime.now().strftime("%H:%M:%S")] + list(tuples_clients[0]))
            if len(tuples_clients):
                cursor.executemany(self.leSQLcl.text(),tuples_clients)
                dbconn.commit()
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'update clients отработал'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'clients - нет запроса'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'clients - нет данных'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'clients - запрос не исполнен'])
        if self.leSQLco.text():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), self.leSQLco.text()])
            ws_log.append([datetime.now().strftime("%H:%M:%S")] + list(tuples_contracts[0]))
            if len(tuples_contracts):
                cursor.executemany(self.leSQLco.text(),tuples_contracts)
                dbconn.commit()
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'update contracts отработал'])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'contracts - нет запроса'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'contracts - нет данных'])
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'contracts - запрос не исполнен'])
        wb_log.save(log_name)
        q=0

    def click_pbDelDoubles(self):
        if not self.file_touched:                               # Проверяем достаточность данных
            self.frFile.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leAgent.isEnabled() and not self.agent_touched:
            self.frAgent.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leFond.isEnabled() and not self.fond_touched:
            self.frFond.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leSigner.isEnabled() and not self.signer_touched:
            self.frSigner.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return

        # Создаем файл с исходными данными и логом
        wb_log = openpyxl.Workbook(write_only=True)

        ws_log = wb_log.create_sheet('Лог')
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Начинаем'])

        log_name = DIR4DELDOUBLESPHONES + datetime.now().strftime('%Y-%m-%d_%H-%M')
        if self.fond_touched:
            log_name += 'ф' + str(self.fond_ids[self.cmbFond.currentIndex()])
        if self.agent_touched:
            log_name += 'а' + str(self.agent_ids[self.cmbAgent.currentIndex()])
        log_name += '.xlsx'

        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Состояние программы:'])
        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'файл ', self.file_name])
        if self.leFond.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', self.cmbFond.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', 'не выбран'])
        if self.leAgent.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', self.cmbAgent.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', 'не выбран'])
        if self.leSigner.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', self.cmbSigner.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', 'не выбран'])

        # Список телефонов у партнера в фонде в который переносим
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        cursor.execute('SELECT partner_code FROM offices_staff WHERE code = %s',
                       (self.agent_ids[self.cmbAgent.currentIndex()],))
        partner = cursor.fetchall()
        if self.partner != partner[0][0]:
            self.partner = partner[0][0]
            phones = []
            cursor = dbconn.cursor()
            sql_tel = 'SELECT ca.client_phone FROM saturn_crm.clients AS cl ' \
                      'LEFT JOIN offices_staff AS os ON cl.inserted_user_code = os.code ' \
                      'LEFT JOIN contracts AS co ON co.client_id = cl.client_id ' \
                      'LEFT JOIN callcenter AS ca ON ca.contract_id = co.id ' \
                      'WHERE os.partner_code = %s GROUP BY ca.client_phone'
            cursor.execute(sql_tel, (self.partner,))
            phones_sql = cursor.fetchall()
            self.progressBar.setMaximum(len(phones_sql) - 1)
            for i, phone_sql in enumerate(phones_sql):
                if not (i % 100):
                    self.progressBar.setValue(i)
                if phone_sql[0] and fine_phone(phone_sql[0]) not in phones:
                    phones.append(fine_phone(phone_sql[0]))
            self.phones = phones
        sheet = self.wb[self.wb.sheetnames[self.cmbTab.currentIndex()]]
        if not sheet.max_row:
            self.errMessage('Файл Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
            return
        table = []
        table_j_end = 0 # Если больше 10 пустых ячеек - на следующую срочку
        table_k_end = 0 # Если больше 10 пустых строчек - заканчиваем чтение таблицы
        self.progressBar.setMaximum(sheet.max_row - 1)
        for j, row in enumerate(sheet.rows):
            if not (j % 100):
                self.progressBar.setValue(j)
            if table_j_end == 10 and j == 10:
                break
            table.append([])
            for k, cell in enumerate(row):
                table[j].append(cell.value)
                if cell.value != None:
                    table_j_end = 0
                    table_k_end = 0
                else:
                    table_j_end += 1
                    table_k_end += 1
                if table_k_end > 10:
                    break

        if not self.chbNoBackup.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Дублируем исходную excel таблицу в этот файл'])
            ws_input = wb_log.create_sheet('Исходная таблица')
            for table_row in table:
                row = []
                for cell in table_row:
                    row.append(cell)
                ws_input.append(row)
        # Удаляем значение в ячейке если это телефон и он есть у партнера
        table_rez = []
        self.progressBar.setMaximum(len(table) - 1)
        for j, row in enumerate(table):
            if not (j % 100):
                self.progressBar.setValue(j)
            table_rez.append([])
            for cell in row:
                if len(str(l(cell))) < 10 or len(str(l(cell))) > 11:
                    table_rez[j].append(cell)
                else:
                    if fine_phone(l(cell)) in self.phones:
                        table_rez[j].append('')
                    else:
                        table_rez[j].append(str(l(fine_phone(cell))))
        self.progressBar.setValue(len(table) - 1)
        ws_rez = wb_log.create_sheet('Без телефонных дублей')
        for row in table_rez:
            ws_rez.append(row)
        wb_log.save(log_name)

    def change_leAgent(self):
        if self.agent_touched:
            agent_id = self.agent_ids[self.cmbAgent.currentIndex()]
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        if self.leAgent.text().strip():
            sql = "SELECT CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, '-', position_id), code " \
                  "FROM offices_staff " \
                  "WHERE CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, user_lastname) LIKE %s " #\
                  #"AND user_fired = 0 "
            cursor.execute(sql, ('%' + self.leAgent.text() + '%',))
        else:
            sql = "SELECT CONCAT_WS(' ', code, '-', user_surname, user_name, user_lastname, '-', position_id), code " \
                  "FROM offices_staff " #\
                  #"WHERE user_fired = 0"
            cursor.execute(sql)
        rows = cursor.fetchall()
        agents = []
        self.agent_names = {}
        self.agent_ids = []
        for i, row in enumerate(rows):
            agents.append(row[0])
            self.agent_names[row[1]] = row[0]
            self.agent_ids.append(row[1])
        self.cmbAgent.clear()
        self.cmbAgent.addItems(agents)
        try:
            if self.agent_touched:
                self.cmbAgent.setCurrentIndex(self.agent_ids.index(agent_id))
        except ValueError:
            self.agent_touched = False

    def click_pbAgent(self):
        self.frAgent.setStyleSheet("QFrame{background-image: }")
        self.leAgent.setEnabled(not self.leAgent.isEnabled())
        self.cmbAgent.setEnabled(not self.cmbAgent.isEnabled())
        return

    def click_pbFond(self):
        self.frFond.setStyleSheet("QFrame{background-image: }")
        self.leFond.setEnabled(not self.leFond.isEnabled())
        self.cmbFond.setEnabled(not self.cmbFond.isEnabled())
        return

    def click_pbSigner(self):
        self.frSigner.setStyleSheet("QFrame{background-image: }")
        self.leSigner.setEnabled(not self.leSigner.isEnabled())
        self.cmbSigner.setEnabled(not self.cmbSigner.isEnabled())
        return

    def set_cmbAgent(self):
        self.agent_touched = True
        return

    def set_cmbFond(self):
        self.fond_touched = True
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        if self.leSigner.text().strip():
            if self.fond_touched:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers " \
                      "WHERE CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname) LIKE %s " \
                      "AND subdomain_id = %s"
                cursor.execute(sql, (self.fond_ids[self.cmbFond.currentIndex()],))
            else:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers " \
                      "WHERE CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname) LIKE %s " \
                      "AND subdomain_id IN (2,6,8,11,12,13,14) AND subdomain_id IN (" + self.fonds_str + ")"
                cursor.execute(sql, ('%' + self.leSigner.text() + '%',))
        else:
            if self.fond_touched:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers WHERE subdomain_id = %s"
                cursor.execute(sql, (self.fond_ids[self.cmbFond.currentIndex()],))
            else:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers WHERE subdomain_id IN (" + self.fonds_str + ") " \
                      "AND subdomain_id IN (2,6,8,11,12,13,14)"
                cursor.execute(sql)
        rows = cursor.fetchall()
        self.signer_ids = []
        self.signer_names = {}
        signers = []
        for i, row in enumerate(rows):
            self.signer_names[row[1]] = row[0]
            self.signer_ids.append(row[1])
            signers.append(row[0])
        self.cmbSigner.clear()
        self.cmbSigner.addItems(signers)
        return

    def set_cmbSigner(self):
        self.signer_touched = True
        return

    def set_cmbCfgFile(self):
        self.frCfgFile.setStyleSheet("QFrame{background-image: }")
        self.cfg_file_touched = True
        self.cfg_file_loaded = False

    def click_clbXAgent(self):
        self.agent_touched = False
        self.leAgent.setText('')
        return

    def click_clbXSigner(self):
        self.signer_touched = False
        self.leSigner.setText('')
        return

    def click_clbXFond(self):
        self.fond_touched = False
        self.leFond.setText('')
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        if self.leSigner.text().strip():
            if self.fond_touched:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers " \
                      "WHERE CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname) LIKE %s " \
                      "AND subdomain_id = %s"
                cursor.execute(sql, (self.fond_ids[self.cmbFond.currentIndex()],))
            else:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers " \
                      "WHERE CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname) LIKE %s " \
                      "AND subdomain_id IN (2,6,8,11,12,13,14) AND subdomain_id IN (" + self.fonds_str + ")"
                cursor.execute(sql, ('%' + self.leSigner.text() + '%',))
        else:
            if self.fond_touched:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers WHERE subdomain_id = %s"
                cursor.execute(sql, (self.fond_ids[self.cmbFond.currentIndex()],))
            else:
                sql = "SELECT CONCAT_WS(' ', id, '-', signer_surname, signer_name, signer_lastname), id " \
                      "FROM signers WHERE subdomain_id IN (" + self.fonds_str + ") " \
                      "AND subdomain_id IN (2,6,8,11,12,13,14)"
                cursor.execute(sql)
        rows = cursor.fetchall()
        self.signer_ids = []
        self.signer_names = {}
        signers = []
        for i, row in enumerate(rows):
            self.signer_names[row[1]] = row[0]
            self.signer_ids.append(row[1])
            signers.append(row[0])
        self.cmbSigner.clear()
        self.cmbSigner.addItems(signers)
        return

    def change_leDir(self):
        try:
            files = sorted(os.listdir(path=self.leDir.text()))
        except OSError:
            return
        self.cmbFile.clear()
        self.cmbFile.addItems(files)
        self.twAllExcels.setColumnCount(0)
        self.twAllExcels.setRowCount(0)
        self.file_touched = False
        self.file_loaded = False
        self.file_name = ''
        return

    def click_chbDateFrom(self):
        if self.deDateFrom.isEnabled():
            self.deDateFrom.setEnabled(False)
        else:
            self.deDateFrom.setEnabled(True)

    def click_chbDateTo(self):
        if self.deDateTo.isEnabled():
            self.deDateTo.setEnabled(False)
        else:
            self.deDateTo.setEnabled(True)


    def set_cmbFile(self):
        if self.leDir.text()[len(self.leDir.text()) - 1:] != '/':
            self.leDir.setText(self.leDir.text() + '/')
        if self.cmbFile.currentText()[len(self.cmbFile.currentText()) - 5:] == '.xlsx':
            self.wb = openpyxl.load_workbook(filename=self.leDir.text() + self.cmbFile.currentText(),
                                        read_only=True)
            self.file_name = self.leDir.text() + self.cmbFile.currentText()
            self.cmbTab.clear()
            self.cmbTab.addItems(self.wb.sheetnames)
        else:
            return

    def set_cmbTab(self):
        self.frFile.setStyleSheet("QFrame{background-image: }")
        self.file_touched = True
        if self.MoveImportPasport == 1:
            self.load4move()
        elif self.MoveImportPasport == 2:
            self.load4import()
        return

    def selectAction(self):
        if self.MoveImportPasport == 1:
            self.frMove.show()
            self.frMoveInf.show()
            self.frImport.hide()
            self.frImportInf.hide()
            self.frPasport.hide()
            self.frPasportInf.hide()
            self.frDelDoublesPhones.hide()
            self.twParsingResult.hide()
        elif self.MoveImportPasport == 2:
            self.frImport.show()
            self.frImportInf.show()
            self.frMove.hide()
            self.frMoveInf.hide()
            self.frPasport.hide()
            self.frPasportInf.hide()
            self.frDelDoublesPhones.hide()
            self.twParsingResult.show()
        elif self.MoveImportPasport == 4:
            self.frDelDoublesPhones.show()
            self.frImportInf.hide()
            self.frMove.hide()
            self.frMoveInf.hide()
            self.frPasport.hide()
            self.frPasportInf.hide()
            self.twParsingResult.hide()
        else:
            self.frPasport.show()
            self.frPasportInf.show()
            self.frMove.hide()
            self.frMoveInf.hide()
            self.frImport.hide()
            self.frImportInf.hide()
            self.frDelDoublesPhones.hide()
            self.twParsingResult.hide()

    def click_clbMove(self):
        self.MoveImportPasport = 2
        self.selectAction()

    def click_clbImport(self):
        self.MoveImportPasport = 3
        self.selectAction()

    def click_clbPasport(self):
        self.MoveImportPasport = 4
        self.selectAction()

    def click_clbDelDoublesPhones(self):
        self.MoveImportPasport = 1
        self.selectAction()


    def load4move(self):
        self.sheet = self.wb[self.wb.sheetnames[self.cmbTab.currentIndex()]]
        if not self.sheet.max_row:
            self.errMessage('Файл Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
            return
        keys = {}
        last_cell = 0
        for j, row in enumerate(self.sheet.rows):
            if j == 0:
                for k, cell in enumerate(row):  # Проверяем, чтобы был СНИЛС
                    if str(cell.value).upper() in IN_IDS:
                        keys[IN_IDS[0]] = k
                if len(keys) > 0:
                    for k, cell in enumerate(row):
                        for n, name in enumerate(IN_NAMES):
                            if n == 0:
                                continue
                            if cell.value != None:
                                if str(cell.value).strip() != '':
                                    last_cell = k
                                    if str(cell.value).upper() == name:
                                        keys[name] = k

                else:
                    self.errMessage('В файле ' + self.cmbFile.currentText() + ' на вкладке ' + self.cmbTab.currentText() +
                                    ' отсутствует колонка с ID')
                    return
            elif j == 1:
                for k, cell in enumerate(row):
                    if cell.value != None:
                        if str(cell.value).strip() != '':
                            if k > last_cell:
                                last_cell = k

        self.twAllExcels.setColumnCount(len(keys))                 # Устанавливаем кол-во колонок
        self.twAllExcels.setRowCount(len(list(self.sheet.rows)) - 1)   # Кол-во строк из таблицы
        self.clients_ids = []
        for j, row in enumerate(self.sheet.rows):
            if j == 0:
                continue
            for k, key in enumerate(keys):
                self.twAllExcels.setItem(j - 1, k, QTableWidgetItem(str(row[keys[key]].value)))
                if k == 0:
                    self.clients_ids.append(row[keys[key]].value)

        # Устанавливаем заголовки таблицы
        self.twAllExcels.setHorizontalHeaderLabels(list(keys))
        # Устанавливаем выравнивание на заголовки
        self.twAllExcels.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twAllExcels.resizeColumnsToContents()
        self.file_loaded = True
        self.table_loaded = False
        if self.leDir.text()[len(self.leDir.text()) - 1:] != '/':
            self.leDir.setText(self.leDir.text() + '/')
        self.file_name = self.leDir.text() + self.cmbFile.currentText() + "!'" + self.cmbTab.currentText() + "'"
        self.table = []
        for i, row in enumerate(self.sheet.rows):
            table_row = []
            for j, cell in enumerate(row):
                if j > last_cell:
                    break
                table_row.append(cell.value)
            self.table.append(table_row)
        return

# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! ПРОВЕРКА ПАСПОРТА !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    def click_pbIdGenerate4PasportCheck(self):
        self.refresh()
        self.frPasport.setStyleSheet("QFrame{background-image: }")
                                                                                # Формируем запрос
        sql_cl = 'SELECT client_id, p_seria, p_number, number, ' \
                 'p_surname, p_name, p_lastname FROM clients AS cl WHERE'
        if self.leAgent.isEnabled():
            sql_cl += ' cl.inserted_user_code = %s'
        if self.leFond.isEnabled():
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ' AND'
            sql_cl += ' cl.subdomain_id = %s'
        if self.chbDateFrom.isChecked():
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ' AND'
            sql_cl += ' cl.inserted_date >= %s'
        if self.chbDateTo.isChecked():
            if sql_cl[len(sql_cl) - 2:] == '%s':
                sql_cl += ' AND'
            sql_cl += ' cl.inserted_date <= %s'
        if sql_cl[len(sql_cl) - 5:] == 'WHERE':
            self.errMessage('Нет ни одного ЗЛ')
            return
        tuple_client = tuple()                                  # Формируем переменные для запросов
        if self.leAgent.isEnabled():
            tuple_client += (self.agent_ids[self.cmbAgent.currentIndex()],)
        if self.leFond.isEnabled():
            tuple_client += (self.fond_ids[self.cmbFond.currentIndex()],)
        if self.chbDateFrom.isChecked():
            tuple_client += (self.deDateFrom.dateTime().toPyDateTime(),)
        if self.chbDateTo.isChecked():
            tuple_client += (self.deDateTo.dateTime().toPyDateTime() + timedelta(hours=23,minutes=59),)
        dbconn = MySQLConnection(**self.dbconfig)
        cursor = dbconn.cursor()
        cursor.execute(sql_cl, tuple_client)
        rows = cursor.fetchall()

        if len(rows):
            self.twAllExcels.setColumnCount(len(rows[0]))                 # Устанавливаем кол-во колонок
            self.twAllExcels.setRowCount(len(rows))   # Кол-во строк из таблицы
            self.clients_ids = []
            for j, row in enumerate(rows):
                for k, cell in enumerate(row):
                    self.twAllExcels.setItem(j, k, QTableWidgetItem(str(cell)))
                    if k == 0:
                        self.clients_ids.append(str(cell))

            # Устанавливаем заголовки таблицы
            self.twAllExcels.setHorizontalHeaderLabels(cursor.column_names)
            # Устанавливаем выравнивание на заголовки
            self.twAllExcels.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
            # делаем ресайз колонок по содержимому
            self.twAllExcels.resizeColumnsToContents()
            self.table = rows
            self.table_loaded = True
            self.file_loaded = False
        else:
            self.errMessage('Нет ЗЛ по данному запросу')

    def click_pbPasportCheck(self):
        if self.leAgent.isEnabled() and not self.agent_touched: # Проверяем достаточность данных
            self.frAgent.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leFond.isEnabled() and not self.fond_touched:
            self.frFond.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if not self.table_loaded:
            self.frPasport.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
                                                                # Создаем файл с исходными данными и логом
        wb_log = openpyxl.Workbook(write_only=True)

        ws_log = wb_log.create_sheet('Лог')
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Начинаем'])

        log_name = DIR4PCHECK + datetime.now().strftime('%Y-%m-%d_%H-%M')
        if self.fond_touched:
            log_name += 'ф' + str(self.fond_ids[self.cmbFond.currentIndex()])
        if self.agent_touched:
            log_name += 'а' + str(self.agent_ids[self.cmbAgent.currentIndex()])
        log_name += '.xlsx'

        #ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Дублируем исходную excel таблицу в этот файл'])
        #ws_input = wb_log.create_sheet('Исходная таблица')
        #for table_row in self.table:
        #    row = []
        #    for cell in table_row:
        #        row.append(cell)
        #    ws_input.append(row)

        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Состояние программы:'])
        if self.leFond.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', self.cmbFond.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', 'не выбран'])
        if self.leAgent.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', self.cmbAgent.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', 'не выбран'])
        if self.chbDateFrom.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Выборка ОТ', self.deDateFrom.date().toPyDate()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Выборка ОТ', 'не выбрана'])
        if self.chbDateTo.isChecked():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Выборка ДО', self.deDateTo.date().toPyDate()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Выборка ДО', 'не выбрана'])

        #ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Бэкап исходного состояния БД создан'])
        #all_clients_ids = "'" + self.clients_ids[0] + "'"       # Проверка на дубли clients
        #for i, client_id in enumerate(self.clients_ids):
        #    if i == 0:
        #        continue
        #    all_clients_ids += ",'" + client_id + "'"
        #ws_backup = wb_log.create_sheet('бэкап БД')
        #dbconn = MySQLConnection(**self.dbconfig)
        #cursor = dbconn.cursor()
        #sql = "SELECT cl.*, co.* FROM clients AS cl LEFT JOIN contracts AS co " \
        #      "ON (cl.client_id = co.client_id) WHERE cl.client_id IN (" + all_clients_ids + ")"
        #cursor.execute(sql)
        #dbrows = cursor.fetchall()
        #ws_backup.append(cursor.column_names)
        #for dbrow in dbrows:
        #    row = []
        #    for dbcell in dbrow:
        #        row.append(dbcell)
        #    ws_backup.append(row)

        p = Popen(['ls', '-l', '--time-style=long-iso', 'list_of_expired_passports.csv.bz2'], stdout=PIPE)
        out, err = p.communicate()
        has_passports = False
        if not p.returncode and out:
            if datetime.now() - timedelta(days=3) < datetime.strptime(out.decode('utf-8').split(' ')[5] +
                                                    ' ' + out.decode('utf-8').split(' ')[6], '%Y-%m-%d %H:%M'):
                has_passports = True

        if not has_passports:
            try:
                os.remove('list_of_expired_passports.csv.bz2')
            except:
                q = 0
            try:
                os.remove('list_of_expired_passports.csv')
            except:
                q = 0

            i = 0                                   # Если база паспортов с ГУМВД устаревшая - скачиваем
            ok = 1
            while ok != 0 and i < 10:
                p = Popen(['wget', '-q', '-c', '-t5',
                           'https://guvm.mvd.ru/upload/expired-passports/list_of_expired_passports.csv.bz2'], stdout=PIPE)
                out, err = p.communicate()
                ok = p.returncode
                i += 1
            if i >= 10:
                print(datetime.now().strftime("%d.%m.%Y %H:%M:%S"), ' Не скачивается, наверное погода нелетная :)')
                return

            all_files = os.listdir(path=".")        # Распаковываем все bzip2 в директории
            for i, all_file in enumerate(all_files):
                if all_file.endswith(".bz2"):
                    with open(all_file.replace('.bz2', ''), 'wb') as new_file, bz2.BZ2File(all_file, 'rb') as file:
                        for data in iter(lambda: file.read(100 * 1024), b''):
                            new_file.write(data)

        has_files = False                       # Проверяем есть ли .csv
        all_files = os.listdir(path=".")
        for all_file in all_files:
            if all_file.endswith(".csv"):
                has_files = True
                new_csv = all_file
        if not has_files:
            print(datetime.now().strftime("%H:%M:%S"), ' В скачанном архиве нет .csv')
            try:
                os.remove('list_of_expired_passports.csv.bz2')
            except:
                q = 0
                return

        if not len(self.passports):
            self.progressBar.setMaximum(118000000)
            with open("list_of_expired_passports.csv","rt") as file_passports:
                for i,line in enumerate(file_passports):
                    if i:
                        if not i%1000000:
                            self.progressBar.setValue(i)
                        self.passports[l(line)] = 1

        self.progressBar.setMaximum(len(self.table)-1)
        ws_pasport = wb_log.create_sheet('Проверка паспортов')
        ws_pasport.append(['ID', 'Серия', 'Номер', 'СНИЛС', 'Фамилия', 'Имя', 'Отчество', 'Проверка паспорта'])  # добавляем первую строку xlsx
        dbconn_saturn = MySQLConnection(**self.dbconfig)
        bad_passport_ids = []
        for j, row in enumerate(self.table):                            # Проверяем паспорта из таблицы
            check = l(row[1])*1000000 + l(row[2])
            try:
                q = self.passports[check]
                rez = 'плохой'
                if self.chbSetStatusInSaturn.isChecked():
                    bad_passport_ids.append((row[0],))
            except KeyError:
                rez = 'OK'

#            if check in self.passports:
#                rez = 'плохой'
#                if self.chbSetStatusInSaturn.isChecked():
#                    bad_passport_ids.append((row[0],))
#            else:
#                rez = 'OK'
#--------------------------------------------------------------------------
#            try:
#                q = self.passports.index(l(row[1])*1000000 + l(row[2]))
#                rez = 'плохой'
#                if self.chbSetStatusInSaturn.isChecked():
#                    bad_passport_ids.append((row[0],))
#            except ValueError:
#                rez = 'OK'
#---------------------------------------------------------------------------
#            for passport in self.passports:
#                if l(row[1]) == l(passport)// 1000000 and l(row[2]) == l(passport)  % 1000000:
#                    rez = 'плохой'
#                    if self.chbSetStatusInSaturn.isChecked():
#                        bad_passport_ids.append((row[0],))
#                    break
            ws_pasport.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], rez])
            if not j%100:
                self.progressBar.setValue(j)
            if len(bad_passport_ids) and not len(bad_passport_ids)%1000:
                write_cursor = dbconn_saturn.cursor()
                write_cursor.executemany('UPDATE contracts AS co SET co.status_secure_code = 6 WHERE co.client_id = %s',
                                         bad_passport_ids)
                dbconn_saturn.commit()
                bad_passport_ids = []
        if len(bad_passport_ids):
            write_cursor = dbconn_saturn.cursor()
            write_cursor.executemany('UPDATE contracts AS co SET co.status_secure_code = 6 WHERE co.client_id = %s',
                                     bad_passport_ids)
            dbconn_saturn.commit()
        self.progressBar.setValue(0)



        qq = """
        self.progressBar.setMaximum(len(self.table)-1)
        ws_pasport = wb_log.create_sheet('Проверка паспортов')
        ws_pasport.append(['ID', 'Серия', 'Номер', 'СНИЛС', 'Фамилия', 'Имя', 'Отчество', 'Проверка паспорта'])  # добавляем первую строку xlsx
        dbconn_pasp = MySQLConnection(**self.dbconfig_pasp)
        dbconn_saturn = MySQLConnection(**self.dbconfig)
        for j, row in enumerate(self.table):                            # Проверяем паспорта из таблицы
            rez = 'OK'
            read_cursor = dbconn_pasp.cursor()
            read_cursor.execute('SELECT p_seria, p_number FROM passport_greylist WHERE p_seria = %s AND p_number = %s',
                                (l(row[1]), l(row[2])))
            row_msg = read_cursor.fetchall()
            if len(row_msg) > 0:
                if self.chbSetStatusInSaturn.isChecked():
                    write_cursor = dbconn_saturn.cursor()
                    write_cursor.execute('UPDATE contracts AS co SET co.status_secure_code = 6 WHERE co.client_id = %s',
                                         (row[0],))
                rez = 'плохой'
            else:
                rez = 'ОК'
            ws_pasport.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], rez])
            self.progressBar.setValue(j)
        dbconn_saturn.commit()
        """

        wb_log.save(log_name)

    #!!!!!!!!!!!!!!!!!!!!! IMPORT !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    def click_pbSaveCfgFile(self):
        wb_cfg = openpyxl.Workbook(write_only=True)
        ws_cfg = wb_cfg.create_sheet('Конфиг')
        for i in range(self.tableWidget.rowCount()):
            ws_cfg.append([self.tableWidget.cellWidget(i,0).currentIndex(), self.tableWidget.cellWidget(i,1).currentIndex()])
        if self.file_loaded:
            wb_cfg.save(DIR4CFGIMPORT + self.cmbFile.currentText() + '.xlsx')
        rows = sorted(os.listdir(DIR4CFGIMPORT))                        # обновляем список конфигов
        cfg_files = []
        for row in rows:
            if row.find('.xlsx') > -1:
                cfg_files.append(row)
        self.cfg_file_names = {}
        for i, cfg_file in enumerate(cfg_files):
            self.cfg_file_names[cfg_file] = i
        self.cmbCfgFile.clear()
        self.cmbCfgFile.addItems(cfg_files)


    def click_pbRefreshImport(self):
        self.refresh()
        self.load4import()

    def updateProgressBar(self, val):
        self.progressBar.setValue(val)

    def load4import(self):  # Заполняем табличку
        if not self.file_touched:                               # Проверяем достаточность данных
            self.frFile.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if not self.cfg_file_touched:
            self.frCfgFile.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return

        self.progressBar.setMaximum(0)
        self.sheet = self.wb[self.wb.sheetnames[self.cmbTab.currentIndex()]]
        if not self.sheet.max_row:
            self.errMessage('Файл Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
            return
        head = []
        for i, row in enumerate(self.sheet.rows):
            if i == 0:
                for cell in row:
                    head.append(cell.value)
                break
        self.progressBar.setMaximum(self.sheet.max_row - 1)

        if not self.cfg_file_loaded:
            for j in range(self.tableWidget.rowCount()):
                self.tableWidget.removeRow(0)
            conf_mass = []
            mass = []
            wb_conf = openpyxl.load_workbook(filename=DIR4CFGIMPORT + self.cmbCfgFile.currentText(),
                                             read_only=True)
            sheet_conf = wb_conf[wb_conf.sheetnames[0]]
            try:
                for row in sheet_conf:
                    conf_mas = []
                    for j, cell in enumerate(row):
                        if j < 2:
                            conf_mas.append(int(cell.value))
                    conf_mass.append(conf_mas)
            except ValueError:
                self.cfg_file_loaded = False
            else:
                self.cfg_file_loaded = True

            for i in FIELDS_IN_RESULT_TABLE_SHORT:
                for name in i:
                    mass.append(name)

            for i, j in enumerate(mass):
                if i < len(conf_mass) and self.cfg_file_loaded:
                    #                if len(conf_mass[i]) < len(mass):
                    if len(conf_mass[i]) > 1:
                        combo1index = conf_mass[i][0]
                        combo2index = conf_mass[i][1]
                    else:
                        combo1index = i  # по умолчанию
                        combo2index = i
                else:
                    combo1index = i  # по умолчанию
                    combo2index = i
                                                        # Заполняем комбобоксы в таблице
                try:  # combo2index - индекс выбора второго QCombobox
                    if head is None:
                        pass
                except AttributeError:
                    return
                self.tableWidget.insertRow(self.tableWidget.rowCount().real)
                items = []

                self.combobox_table_result = QComboBox()
                # self.combobox_table_result.setMaxVisibleItems(15)
                for row in FIELDS_IN_RESULT_TABLE_FULL:
                    for name in row:
                        self.combobox_table_result.addItem(name)
                items.append(self.combobox_table_result)
                if combo1index != -1:
                    self.combobox_table_result.setCurrentIndex(combo1index)  # combobox_table_result - первый комбобокс
                name_combobox_table_result = "combobox_table_result_{0}".format(self.tableWidget.rowCount() - 1)
                self.combobox_table_result.setObjectName(name_combobox_table_result)

                self.combobox_table_from = QComboBox()
                # self.combobox_table_from.setMaxVisibleItems(15)
                for name in head:
                    name = str(name)
                    self.combobox_table_from.addItem(name)
                items.append(self.combobox_table_from)
                if combo2index != -1:
                    self.combobox_table_from.setCurrentIndex(combo2index)  # combobox_table_from - второй комбобокс
                name_combobox_table_from = "combobox_table_from_{0}".format(self.tableWidget.rowCount() - 1)
                self.combobox_table_from.setObjectName(name_combobox_table_from)

                for n, i in enumerate(items):
                    self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, n, i)
            self.tableWidget.resizeColumnsToContents()

        keys = {}
        last_cell = 0
        self.has_gen_snils = False
        for num_item in range(self.tableWidget.rowCount()):
            if self.tableWidget.cellWidget(num_item, 0).currentText() == 'Генератор некорректных СНИЛС':
                self.has_gen_snils = True
        for j, row in enumerate(self.sheet.rows):
            if j == 0:
                for k, cell in enumerate(row):  # Проверяем, чтобы был СНИЛС
                    if str(cell.value).strip().upper() in IN_SNILS:
                        keys[IN_SNILS[0]] = k
                if self.has_gen_snils:
                    q=0
                elif len(keys) > 0:
                    for k, cell in enumerate(row):
                        for n, name in enumerate(IN_NAMES):
                            if n == 0:
                                continue
                            if cell.value != None:
                                if str(cell.value).strip() != '':
                                    last_cell = k
                                    if str(cell.value).upper() == name:
                                        keys[name] = k
                else:
                    self.errMessage('В файле ' + self.cmbFile.currentText() + ' на вкладке ' + self.cmbTab.currentText() +
                                    ' отсутствует колонка со СНИЛС')
                    return
            elif j == 1:
                for k, cell in enumerate(row):
                    if cell.value != None:
                        if str(cell.value).strip() != '':
                            if k > last_cell:
                                last_cell = k
        if not self.has_gen_snils:
            self.clients_snils = []                                         # Добавляем СНИЛСы (для проверки на дубли)
            for i, row in enumerate(self.sheet):
                if i == 1:
                    continue
                self.clients_snils.append(l(row[keys[IN_SNILS[0]]].value))

        self.twAllExcels.setColumnCount(last_cell + 1)                      # Отображаем исходную таблицу
        self.twAllExcels.setRowCount(len(list(self.sheet.rows)) - 1)
        self.clients_ids = []
        headers = []
        for j, row in enumerate(self.sheet.rows):
            for k, cell in enumerate(row):
                if j == 0:
                    headers.append(s(cell.value))
                else:
                    self.twAllExcels.setItem(j - 1, k, QTableWidgetItem(s(cell.value)))

        self.twAllExcels.setHorizontalHeaderLabels(headers)             # Устанавливаем заголовки таблицы
        # Устанавливаем выравнивание на заголовки
        self.twAllExcels.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twAllExcels.resizeColumnsToContents()
        self.file_loaded = True
        self.table_loaded = False

        for i, row in enumerate(self.sheet.rows):
            table_row = []
            for j, cell in enumerate(row):
                if j > last_cell:
                    break
                table_row.append(cell.value)
            self.table.append(table_row)
        self.previewImport()

    def errMessage(self, err_text):  ## Method to open a message box
        infoBox = QMessageBox()  ##Message Box that doesn't run
        infoBox.setIcon(QMessageBox.Warning)
        infoBox.setText(err_text)
        #        infoBox.setInformativeText("Informative Text")
        infoBox.setWindowTitle(datetime.strftime(datetime.now(), "%H:%M:%S") + ' Ошибка: ')
        #        infoBox.setDetailedText("Detailed Text")
        #        infoBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        infoBox.setStandardButtons(QMessageBox.Ok)
        #        infoBox.setEscapeButton(QMessageBox.Close)
        infoBox.exec_()


    def click_pbImport(self):
        if not self.file_loaded:                               # Проверяем достаточность данных
            self.frFile.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if not self.cfg_file_loaded:
            self.frCfgFile.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leAgent.isEnabled() and not self.agent_touched:
            self.frAgent.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leFond.isEnabled() and not self.fond_touched:
            self.frFond.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return
        if self.leSigner.isEnabled() and not self.signer_touched:
            self.frSigner.setStyleSheet("QFrame{background-image: url(./x.png)}")
            return

        if self.cmbGenderType.currentIndex() == 0:
            female_gender_value = 'Ж'
            male_gender_value = 'М'
            gender_length = 1
        elif self.cmbGenderType.currentIndex() == 1:
            female_gender_value = '1'
            male_gender_value = '0'
            gender_length = 1
        else:
            female_gender_value = 'Ж'
            male_gender_value = 'М'
            gender_length = 1

            # Создаем директорию
        if datetime.now().strftime("%Y-%m-%d") not in os.listdir(DIR4IMPORT):
            os.mkdir(DIR4IMPORT + datetime.now().strftime("%Y-%m-%d"))
        dir_name = DIR4IMPORT + datetime.now().strftime("%Y-%m-%d") + '/' + \
                   datetime.now().strftime('%H-%M_')
        if self.fond_touched:
            dir_name += 'ф' + str(self.fond_ids[self.cmbFond.currentIndex()])
        if self.agent_touched:
            dir_name += 'а' + str(self.agent_ids[self.cmbAgent.currentIndex()])
        dir_name += '/'
        os.mkdir(dir_name)
        self.file_name = dir_name + self.cmbFile.currentText()
                                                                        # Создаем файл с исходными данными и логом
        log_name = dir_name + self.cmbFile.currentText() + '_' + self.cmbTab.currentText() + '.xlsx'
        wb_log = openpyxl.Workbook(write_only=True)
        ws_log = wb_log.create_sheet('Лог')
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Начинаем'])

        if not self.has_gen_snils:                                      # Проверка на дубли исходной таблицы
            doubles_in_input = list(set([x for x in self.clients_snils if self.clients_snils.count(x) > 1]))
            if len(doubles_in_input):
                ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Дубли в СНИЛС в исходной таблице'])
                ws_input_doubles = wb_log.create_sheet('Дубли в СНИЛС в исходной таблице')
                ws_input_doubles.append(['ID'])
                for row in doubles_in_input:
                    ws_input_doubles.append([normalize_snils(row)])
        ws_log.append([datetime.now().strftime("%H:%M:%S"), ' Состояние программы:'])
        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Исходный файл ', self.file_name])
        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Конфигурационный файл ', self.cmbCfgFile.currentText()])
        if self.leFond.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', self.cmbFond.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'фонд', 'не выбран'])
        if self.leAgent.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', self.cmbAgent.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'агент', 'не выбран'])
        if self.leSigner.isEnabled():
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', self.cmbSigner.currentText()])
        else:
            ws_log.append([datetime.now().strftime("%H:%M:%S"), 'подписант', 'не выбран'])

        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Дублируем исходную excel таблицу в файл лога'])
        ws_input = wb_log.create_sheet('Исходная таблица')
        for table_row in self.table:
            row = []
            for cell in table_row:
                row.append(cell)
            ws_input.append(row)

        self.updateProgressBar(0)
        self.pbImport.setEnabled(False)

        self.workerThread = WorkerThread(sheet=self.sheet, tableWidget=self.tableWidget,
                                         fname=self.file_name, agent=self.agent_ids[self.cmbAgent.currentIndex()],
                                         signer=self.signer_ids[self.cmbSigner.currentIndex()],
                                         cmbGenderType=self.cmbGenderType, cmbParsingType=self.cmbParsingType)
                                                                        # <<<<<<<<<<<<<<<<<<<<<<<<<запускаем подпроцесс
        self.workerThread.progress_value.connect(self.updateProgressBar)
        self.workerThread.start()
        self.updateProgressBar(0)
        self.pbImport.setEnabled(True)
        ws_log.append([datetime.now().strftime("%H:%M:%S"), 'Импорт отработал, файл(ы) создан(ы)'])
        wb_log.save(log_name)


    def previewImport(self):
        if self.cmbGenderType.currentIndex() == 0:
            female_gender_value = 'Ж'
            male_gender_value = 'М'
            gender_length = 1
        elif self.cmbGenderType.currentIndex() == 1:
            female_gender_value = '1'
            male_gender_value = '0'
            gender_length = 1
        else:
            female_gender_value = 'Ж'
            male_gender_value = 'М'
            gender_length = 1
        err_from_log = {}
        self.twParsingResult.setColumnCount(0)
        self.twParsingResult.setRowCount(0)
        self.twParsingResult.setColumnCount(len(HEAD_RESULT_EXCEL_FILE))
        self.twParsingResult.setHorizontalHeaderLabels(HEAD_RESULT_EXCEL_FILE)  # добавляем заголовки
        self.twParsingResult.setRowCount(3)
        maxParsingResult = 0
        for num_row, row in enumerate(self.sheet.rows):
            if num_row == 0:
                continue
            result_row = {}
            passport = Passport()
            phone = Phone()

            for num_item in range(self.tableWidget.rowCount()):
                item0 = self.tableWidget.cellWidget(num_item, 0).currentIndex()
                item1 = self.tableWidget.cellWidget(num_item, 1).currentIndex()
                label0 = self.tableWidget.cellWidget(num_item, 0).currentText()
                label1 = self.tableWidget.cellWidget(num_item, 1).currentText()

                row_item = str(row[item1].value)  # Если преобразовывать все в стринг, то только тут
                if row_item == 'None' or row_item == '2001-01-00' or row_item == '2001-01-00 00:00:00' \
                        or row_item == 'null' or row_item == 'NULL' \
                        or row_item == '\\N' or row_item == '\\n' \
                        or row_item == 'заполнить' or row_item == '00.00.0000' \
                        or row_item == '0000-00-00' or row_item == 'ERROR' \
                        or row_item == '=#ССЫЛ!' or row_item == '#ССЫЛ!' \
                        or row_item == '=#REF!' or row_item == '#REF!' or row_item == '-':
                    row_item = ''
                elif row_item == '0' and label0 != 'Пол':
                    row_item = ''

                if label0 in MANIPULATE_LABELS:

                    if label0 in ["ФИО из поля", "ФИО при рождении из поля"]:
                        FIO = field2fio(row_item)
                        if label0 == "ФИО из поля":
                            lab = FIO_LABELS
                        elif label0 == "ФИО при рождении из поля":
                            lab = FIO_BIRTH_LABELS
                        if row_item == '':
                            for j in range(len(lab)):
                                result_row[lab[j]] = NEW_NULL_VALUE_FOR_ALL_TEXT
                        else:
                            for j in range(len(FIO)):
                                result_row[lab[j]] = FIO[j]
                        continue
                    elif label0 == "Регистрация -> Регион":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[1], ADRESS_REG_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Район":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[3], ADRESS_REG_LABELS[4]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Город":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[5], ADRESS_REG_LABELS[6]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Населенный_пункт":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[7], ADRESS_REG_LABELS[8]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Улица":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[9], ADRESS_REG_LABELS[10]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    # ADRESS_LIVE_LABELS
                    elif label0 == "Проживание -> Регион":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[1], ADRESS_LIVE_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Район":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[3], ADRESS_LIVE_LABELS[4]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Город":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[5], ADRESS_LIVE_LABELS[6]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Населенный_пункт":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[7], ADRESS_LIVE_LABELS[8]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Улица":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[9], ADRESS_LIVE_LABELS[10]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Адрес регистрации из_поля":
                        result_row[ADRESS_REG_LABELS[0]] = '111111'
                        adress_reg = FullAdress(row_item, tip=self.cmbParsingType.currentText())
                        #                        qr = ''
                        for z, cell in enumerate(adress_reg.get_values()):
                            result_row[ADRESS_REG_LABELS[z]] = cell
                        #                            qr += cell + ' '
                        #                        print(qr)
                        n = [char for char in result_row[ADRESS_REG_LABELS[0]] if char in string.digits]
                        if len(n) != 6:
                            result_row[ADRESS_REG_LABELS[0]] = '111111'

                    elif label0 == "Адрес проживания из поля":
                        result_row[ADRESS_LIVE_LABELS[0]] = '111111'
                        adress_zhit = FullAdress(row_item, tip=self.cmbParsingType.currentText())
                        #                        qr = ''
                        for z, cell in enumerate(adress_zhit.get_values()):
                            result_row[ADRESS_LIVE_LABELS[z]] = cell
                        #                            qr += cell + ' '
                        #                        print(qr)
                        n = [char for char in result_row[ADRESS_LIVE_LABELS[0]] if char in string.digits]
                        if len(n) != 6:
                            result_row[ADRESS_LIVE_LABELS[0]] = '111111'

                    elif label0 == "Регион регистрации из номера":
                        addr = field2addr(REGIONS[intl(row_item)])
                        lab = [ADRESS_REG_LABELS[1], ADRESS_REG_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Регион проживания из номера":
                        addr = field2addr(REGIONS[intl(row_item)])
                        lab = [ADRESS_LIVE_LABELS[1], ADRESS_LIVE_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Cерия и Номер паспорта из поля":
                        addr = field2sernum(row_item)
                        lab = [PASSPORT_DATA_LABELS[0], PASSPORT_DATA_LABELS[1]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Генератор некорректных СНИЛС":
                        dbconfig = read_config(filename='move.ini', section='mysql')
                        dbconn = MySQLConnection(**dbconfig)
                        count_snils = 1
                        cached_snils = 0
                        while count_snils > 0:
                            checksum_snils = self.checksum(self.start_snils)
                            for i in range(self.start_snils_cs + 1, 99):
                                if i != checksum_snils:
                                    full_snils = self.start_snils * 100 + i
                                    dbcursor = dbconn.cursor()
                                    dbcursor.execute('SELECT `number` FROM clients WHERE `number` = %s', (full_snils,))
                                    dbchk = dbcursor.fetchall()
                                    if len(dbchk) == 0:
                                        cached_snils = full_snils
                                        count_snils -= 1
                                        self.start_snils_cs = i
                                        break
                            if count_snils > 0:
                                self.start_snils -= 1
                                self.start_snils_cs = 0
                        dbconn.close()
                        result_row[SNILS_LABEL[0]] = normalize_snils(cached_snils)

                elif label0 == '-------------------------':
                    continue
                elif label0 in SNILS_LABEL:
                    if GENERATE_SNILS:
                        dbconfig = read_config(filename='move.ini', section='mysql')
                        dbconn = MySQLConnection(**dbconfig)
                        count_snils = 1
                        cached_snils = 0
                        while count_snils > 0:
                            checksum_snils = self.checksum(self.start_snils)
                            for i in range(self.start_snils_cs + 1, 99):
                                if i != checksum_snils:
                                    full_snils = self.start_snils * 100 + i
                                    dbcursor = dbconn.cursor()
                                    dbcursor.execute('SELECT `number` FROM clients WHERE `number` = %s', (full_snils,))
                                    dbchk = dbcursor.fetchall()
                                    if len(dbchk) == 0:
                                        cached_snils = full_snils
                                        count_snils -= 1
                                        self.start_snils_cs = i
                                        break
                            if count_snils > 0:
                                self.start_snils -= 1
                                self.start_snils_cs = 0
                        dbconn.close()
                        result_row[label0] = normalize_snils(cached_snils)
                    else:
                        result_row[label0] = normalize_snils(row_item)
                elif label0 in PLACE_BIRTH_LABELS:
                    result_row[label0] = row_item
                elif label0 in PASSPORT_DATA_LABELS:
                    if PASSPORT_DATA_LABELS.index(label0) == 0:
                        result_row[label0] = normalize_seria(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 1:
                        result_row[label0] = normalize_index(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 2:
                        result_row[label0] = normalize_date(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 3:
                        result_row[label0] = normalize_text(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 4:
                        result_row[label0] = normalize_index(row_item)

                elif label0 in PHONES_LABELS:
                    if PHONES_LABELS.index(label0) == 0:
                        phone.tel_mob = row_item
                    elif PHONES_LABELS.index(label0) == 1:
                        phone.tel_rod = row_item
                    elif PHONES_LABELS.index(label0) == 2:
                        phone.tel_dom = row_item
                elif label0 in DATE_BIRTH_LABEL:
                    result_row[label0] = normalize_date(row_item)
                elif label0 in GENDER_LABEL:
                    result_row[label0] = normalize_gender(row_item, female_gender_value, male_gender_value, gender_length)
                elif label0 == ADRESS_REG_LABELS[0] or label0 == ADRESS_LIVE_LABELS[0]:
                    result_row[label0] = normalize_index(row_item)
                elif label0 in ADRESS_REG_LABELS[11]:
                    result_row[label0] = normalize_home(row_item)
                elif label0 in ADRESS_LIVE_LABELS[11]:
                    result_row[label0] = normalize_home(row_item)
                elif label0 in ADRESS_REG_LABELS:
                    result_row[label0] = row_item
                elif label0 in ADRESS_LIVE_LABELS:
                    result_row[label0] = row_item
                elif label0 in TECH_LABELS:
                    if label0 == TECH_LABELS[0]:
                        result_row[label0] = self.agent_ids[self.cmbAgent.currentIndex()]
                    elif label0 == TECH_LABELS[1]:
                        result_row[label0] = self.signer_ids[self.cmbSigner.currentIndex()]
                    elif label0 == TECH_LABELS[2]:
                        result_row[label0] = PREDSTRAH_ID
                else:
                    result_row[label0] = normalize_text(row_item)

#            for num, z in enumerate(passport.get_values()):
#                result_row[PASSPORT_DATA_LABELS[num]] = z
            for num, z in enumerate(phone.get_values()):
                result_row[PHONES_LABELS[num]] = z

            LABELS = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, FIO_SNILS_LABELS, GENDER_LABEL, DATE_BIRTH_LABEL,
                      PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS, ADRESS_LIVE_LABELS,
                      PHONES_LABELS, TECH_LABELS]
            mass = []
            for label_group in LABELS:
                for label in label_group:
                    mass.append(label)
            yum = True
            yum_phone0 = -1
            yum_phone1 = -1
            yum_phone2 = -1
            for num, cell in enumerate(mass):
                if cell in result_row:
                    mass[num] = result_row[cell]  # заполняем mass, чтобы его добавить как строку в xlsx
                    if cell == PHONES_LABELS[0]:
                        if mass[num] == ERROR_VALUE:
                            mass[num] = ''
                        yum_phone0 = num
                    elif cell == PHONES_LABELS[1]:
                        if mass[num] == ERROR_VALUE:
                            mass[num] = ''
                        yum_phone1 = num
                    elif cell == PHONES_LABELS[2]:
                        if mass[num] == ERROR_VALUE:
                            mass[num] = ''
                        yum_phone2 = num
                    elif mass[num] == ERROR_VALUE:
                        yum = False
                else:
                    mass[num] = ''  # заполняем mass, чтобы его добавить как строку в xlsx

            if mass[yum_phone0] == mass[yum_phone1] and mass[yum_phone0] != '':  # стираем дублирующиеся телефоны
                mass[yum_phone1] = ''
            if mass[yum_phone1] == mass[yum_phone2] and mass[yum_phone1] != '':
                mass[yum_phone2] = ''
            if mass[yum_phone0] == mass[yum_phone2] and mass[yum_phone0] != '':
                mass[yum_phone2] = ''

            if mass[yum_phone0] == '' and mass[yum_phone1] == '' and mass[yum_phone2] == '':
                yum = False  # если нет ни одного телефона - ошибка

            if yum and err_from_log.get(num_row + 1) == None:
                for ind, cell in enumerate(mass):
                    self.twParsingResult.setItem(maxParsingResult, ind, QTableWidgetItem(str(cell)))
                maxParsingResult += 1
                #ws.append(mass)
                #print(num_row, result_row['ФИО.Фамилия'], result_row['ФИО.Имя'], result_row['ФИО.Отчество'])
            else:
                mass.append(num_row + 1)
                mass.append(err_from_log.get(num_row + 1))
                #ws_err.append(mass)
                print('Ошибка:', num_row, result_row['ФИО.Фамилия'], result_row['ФИО.Имя'], result_row['ФИО.Отчество'])

            if maxParsingResult > 4:
                break
#        self.twParsingResult.horizontalHeaderItem(0).setTextAlignment(Qt.AlignCenter)
        # делаем ресайз колонок по содержимому
        self.twParsingResult.resizeColumnsToContents()

class WorkerThread(QThread):
    progress_value = QtCore.pyqtSignal(int)

    def __init__(self, tableWidget, sheet, fname, agent, signer, parent=None, cmbGenderType=None, cmbParsingType=None):
        super(WorkerThread, self).__init__(parent)
        self.tableWidget = tableWidget
        self.sheet = sheet
        self.fname = fname
        self.agent_id = agent
        self.signer_id = signer
        self.cmbGenderType = cmbGenderType
        self.cmbParsingType = cmbParsingType
        dbconfig = read_config(filename='move.ini', section='mysql')
        dbconn = MySQLConnection(**dbconfig)
        dbcursor = dbconn.cursor()
        dbcursor.execute('SELECT min(`number`) FROM  clients WHERE `number` > 99900000000;')
        dbrows = dbcursor.fetchall()
        dbconn.close()
        self.start_snils = int('{0:011d}'.format(dbrows[0][0])[:-2])  # 9 цифр неправильного СНИЛСа с которого уменьшаем
        self.start_snils_cs = int('{0:011d}'.format(dbrows[0][0])[-2:])  # контрольная сумма неправильного СНИЛСа
        if GENERATE_SNILS:
            self.wb_comp = Workbook(write_only=True)
            self.ws_comp = self.wb_comp.create_sheet('Лист1')
            self.ws_comp.append(['Реальный СНИЛС', 'Псевдо-СНИЛС'])  # добавляем первую строку xlsx

    def checksum(self, snils_dig):  # Вычисляем 2 последних цифры СНИЛС по первым 9-ти
        def snils_csum(sn):
            k = range(9, 0, -1)
            pairs = zip(k, [int(x) for x in sn.replace('-', '').replace(' ', '')])
            return sum([k * v for k, v in pairs])
        snils = '{0:09d}'.format(snils_dig)
        csum = snils_csum(snils)
        while csum > 101:
            csum %= 101
        if csum in (100, 101):
            csum = 0
        return csum

    def run(self):
        self.start_process()

    def start_process(self):
        if self.cmbGenderType.currentIndex() == 0:
            female_gender_value = 'Ж'
            male_gender_value = 'М'
            gender_length = 1
        elif self.cmbGenderType.currentIndex() == 1:
            female_gender_value = '1'
            male_gender_value = '0'
            gender_length = 1
        else:
            female_gender_value = 'Ж'
            male_gender_value = 'М'
            gender_length = 1
        lname = self.fname[0:self.fname.rfind('xlsx')]+ 'log'
        err_from_log = {}
        use_log = False
        try:
            log_file = open(lname,'rt',encoding='utf-8')
            log_file_string = log_file.read()
            first_sq = 0
            next_str = 0
            dub_toch = 0
            last_sq = 1
            n_str_w_err = ''
            text_err = ''
            for nx in range(len(log_file_string)):
                if log_file_string[nx] == ':':
                    dub_toch = nx
                if log_file_string[nx] == '\n':
                    next_str = nx
                if log_file_string[nx] == '#' or nx == len(log_file_string) - 1:
                    first_sq = last_sq
                    last_sq = nx
                    if dub_toch > 0:
                        n_str_w_err = int(log_file_string[first_sq+1:dub_toch])
                        text_err = log_file_string[dub_toch + 3:next_str]
                        err_from_log[n_str_w_err] = text_err
            use_log = True
        except:
            use_log = False

        cname = self.fname[0:self.fname.rfind('.xlsx')]+ '_new.cfg'
        conf_file = open(cname,'wt',encoding='utf-8')
        for i in range(self.tableWidget.rowCount()):
            conf_file.write(str(self.tableWidget.cellWidget(i,0).currentIndex()) + ' ' +
                            str(self.tableWidget.cellWidget(i,1).currentIndex()) + '\n')
        conf_file.close()

        wb_err = Workbook(write_only=True)
        ws_err = wb_err.create_sheet('Ошибки')
        ws_err.append(HEAD_RESULT_EXCEL_FILE)                                         # добавляем первую строку xlsx
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Лист1')
        ws.append(HEAD_RESULT_EXCEL_FILE)                                             # добавляем первую строку xlsx
        dbconfig = read_config(filename='move.ini', section='mysql')
        dbconn = MySQLConnection(**dbconfig)
        dbcursor = dbconn.cursor()

        file_number = 1
        for num_row, row in enumerate(self.sheet.rows):
            self.progress_value.emit(num_row + 1)  # отрисовываем ProgresBar
            if num_row == 0:
                continue
            result_row = {}
            passport = Passport()
            phone = Phone()

            for num_item in range(self.tableWidget.rowCount()):
                item0 = self.tableWidget.cellWidget(num_item, 0).currentIndex()
                item1 = self.tableWidget.cellWidget(num_item, 1).currentIndex()
                label0 = self.tableWidget.cellWidget(num_item, 0).currentText()
                label1 = self.tableWidget.cellWidget(num_item, 1).currentText()

                row_item = str(row[item1].value)                     # Если преобразовывать все в стринг, то только тут
                if row_item == 'None' or row_item == '2001-01-00' or row_item == '2001-01-00 00:00:00' \
                                      or  row_item == 'null' or  row_item == 'NULL' \
                                      or  row_item == '\\N' or  row_item == '\\n' \
                                      or  row_item == 'заполнить' or row_item == '00.00.0000'\
                                      or row_item == '0000-00-00' or row_item == 'ERROR' \
                                      or row_item == '=#ССЫЛ!' or row_item == '#ССЫЛ!'\
                                      or row_item == '=#REF!' or row_item == '#REF!' or row_item == '-':
                    row_item = ''
                elif row_item == '0' and label0 != 'Пол':
                    row_item = ''

                if label0 in MANIPULATE_LABELS:

                    if label0 in ["ФИО из поля", "ФИО при рождении из поля"]:
                        FIO = field2fio(row_item)
                        if label0 == "ФИО из поля":
                            lab = FIO_LABELS
                        elif label0 == "ФИО при рождении из поля":
                            lab = FIO_BIRTH_LABELS
                        if row_item == '':
                            for j in range(len(lab)):
                                result_row[lab[j]] = NEW_NULL_VALUE_FOR_ALL_TEXT
                        else:
                            for j in range(len(FIO)):
                                result_row[lab[j]] = FIO[j]
                        continue

                    elif label0 == "Регистрация -> Регион":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[1], ADRESS_REG_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Район":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[3], ADRESS_REG_LABELS[4]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Город":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[5], ADRESS_REG_LABELS[6]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Населенный_пункт":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[7], ADRESS_REG_LABELS[8]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Регистрация -> Улица":
                        addr = field2addr(row_item)
                        lab = [ADRESS_REG_LABELS[9], ADRESS_REG_LABELS[10]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
# ADRESS_LIVE_LABELS
                    elif label0 == "Проживание -> Регион":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[1], ADRESS_LIVE_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Район":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[3], ADRESS_LIVE_LABELS[4]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Город":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[5], ADRESS_LIVE_LABELS[6]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Населенный_пункт":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[7], ADRESS_LIVE_LABELS[8]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]
                    elif label0 == "Проживание -> Улица":
                        addr = field2addr(row_item)
                        lab = [ADRESS_LIVE_LABELS[9], ADRESS_LIVE_LABELS[10]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Адрес регистрации из_поля":
                        result_row[ADRESS_REG_LABELS[0]] = '111111'
                        adress_reg = FullAdress(row_item, tip=self.cmbParsingType.currentText())
#                        qr = ''
                        for z, cell in enumerate(adress_reg.get_values()):
                            result_row[ADRESS_REG_LABELS[z]] = cell
#                            qr += cell + ' '
#                        print(qr)
                        n = [char for char in result_row[ADRESS_REG_LABELS[0]] if char in string.digits]
                        if len(n) != 6:
                            result_row[ADRESS_REG_LABELS[0]] = '111111'

                    elif label0 == "Адрес проживания из поля":
                        result_row[ADRESS_LIVE_LABELS[0]] = '111111'
                        adress_zhit = FullAdress(row_item, tip=self.cmbParsingType.currentText())
#                        qr = ''
                        for z, cell in enumerate(adress_zhit.get_values()):
                            result_row[ADRESS_LIVE_LABELS[z]] = cell
#                            qr += cell + ' '
#                        print(qr)
                        n = [char for char in result_row[ADRESS_LIVE_LABELS[0]] if char in string.digits]
                        if len(n) != 6:
                            result_row[ADRESS_LIVE_LABELS[0]] = '111111'

                    elif label0 == "Регион регистрации из номера":
                        addr = field2addr(REGIONS[intl(row_item)])
                        lab = [ADRESS_REG_LABELS[1], ADRESS_REG_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Регион проживания из номера":
                        addr = field2addr(REGIONS[intl(row_item)])
                        lab = [ADRESS_LIVE_LABELS[1], ADRESS_LIVE_LABELS[2]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Cерия и Номер паспорта из поля":
                        addr = field2sernum(row_item)
                        lab = [PASSPORT_DATA_LABELS[0],PASSPORT_DATA_LABELS[1]]
                        for j in range(len(addr)):
                            result_row[lab[j]] = addr[j]

                    elif label0 == "Генератор некорректных СНИЛС":
                        count_snils = 1
                        cached_snils = 0
                        while count_snils > 0:
                            checksum_snils = self.checksum(self.start_snils)
                            for i in range(self.start_snils_cs + 1, 99):
                                if i != checksum_snils:
                                    full_snils = self.start_snils * 100 + i
                                    dbcursor.execute('SELECT `number` FROM clients WHERE `number` = %s', (full_snils,))
                                    dbchk = dbcursor.fetchall()
                                    if len(dbchk) == 0:
                                        cached_snils = full_snils
                                        count_snils -= 1
                                        self.start_snils_cs = i
                                        break
                            if count_snils > 0:
                                self.start_snils -= 1
                                self.start_snils_cs = 0
                        result_row[SNILS_LABEL[0]] = normalize_snils(cached_snils)

                elif label0 == '-------------------------':
                    continue
                elif label0 in SNILS_LABEL:
                    if GENERATE_SNILS:
                        dbconfig = read_config(filename='move.ini', section='mysql')
                        dbconn = MySQLConnection(**dbconfig)
                        count_snils = 1
                        cached_snils = 0
                        while count_snils > 0:
                            checksum_snils = self.checksum(self.start_snils)
                            for i in range(self.start_snils_cs + 1, 99):
                                if i != checksum_snils:
                                    full_snils = self.start_snils * 100 + i
                                    dbcursor = dbconn.cursor()
                                    dbcursor.execute('SELECT `number` FROM clients WHERE `number` = %s', (full_snils,))
                                    dbchk = dbcursor.fetchall()
                                    if len(dbchk) == 0:
                                        cached_snils = full_snils
                                        count_snils -= 1
                                        self.start_snils_cs = i
                                        break
                            if count_snils > 0:
                                self.start_snils -= 1
                                self.start_snils_cs = 0
                        dbconn.close()
                        result_row[label0] = normalize_snils(cached_snils)
                        self.ws_comp.append([row_item, cached_snils])
                    else:
                        result_row[label0] = normalize_snils(row_item)
                elif label0 in PLACE_BIRTH_LABELS:
                    result_row[label0] = row_item
                elif label0 in PASSPORT_DATA_LABELS:
                    if PASSPORT_DATA_LABELS.index(label0) == 0:
                        result_row[label0] = normalize_seria(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 1:
                        result_row[label0] = normalize_index(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 2:
                        result_row[label0] = normalize_date(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 3:
                        result_row[label0] = normalize_text(row_item)
                    elif PASSPORT_DATA_LABELS.index(label0) == 4:
                        result_row[label0] = normalize_index(row_item)

                elif label0 in PHONES_LABELS:
                    if PHONES_LABELS.index(label0) == 0:
                        phone.tel_mob = row_item
                    elif PHONES_LABELS.index(label0) == 1:
                        phone.tel_rod = row_item
                    elif PHONES_LABELS.index(label0) == 2:
                        phone.tel_dom = row_item
                elif label0 in DATE_BIRTH_LABEL:
                    result_row[label0] = normalize_date(row_item)
                elif label0 in GENDER_LABEL:
                    result_row[label0] = normalize_gender(row_item, female_gender_value, male_gender_value, gender_length)
                elif label0 == ADRESS_REG_LABELS[0] or label0 == ADRESS_LIVE_LABELS[0]:
                    result_row[label0] = normalize_index(row_item)
                elif label0 in ADRESS_REG_LABELS[11]:
                    result_row[label0] = normalize_home(row_item)
                elif label0 in ADRESS_LIVE_LABELS[11]:
                    result_row[label0] = normalize_home(row_item)
                elif label0 in ADRESS_REG_LABELS:
                    result_row[label0] = row_item
                elif label0 in ADRESS_LIVE_LABELS:
                    result_row[label0] = row_item
                elif label0 in TECH_LABELS:
                    if label0 == TECH_LABELS[0]:
                        result_row[label0] = self.agent_id
                    elif label0 == TECH_LABELS[1]:
                        result_row[label0] = self.signer_id
                    elif label0 == TECH_LABELS[2]:
                        result_row[label0] = PREDSTRAH_ID
                else:
                    result_row[label0] = normalize_text(row_item)

#            for num, z in enumerate(passport.get_values()):
#                result_row[PASSPORT_DATA_LABELS[num]] = z
            for num, z in enumerate(phone.get_values()):
                result_row[PHONES_LABELS[num]] = z

            LABELS = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, FIO_SNILS_LABELS, GENDER_LABEL, DATE_BIRTH_LABEL,
                      PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS, ADRESS_LIVE_LABELS,
                      PHONES_LABELS, TECH_LABELS]
            mass = []
            for label_group in LABELS:
                for label in label_group:
                    mass.append(label)
            yum = True
            yum_phone0 = -1
            yum_phone1 = -1
            yum_phone2 = -1
            for num, cell in enumerate(mass):
                if cell in result_row:
                    mass[num] = result_row[cell]                # заполняем mass, чтобы его добавить как строку в xlsx
                    if cell == PHONES_LABELS[0]:
                        if mass[num] == ERROR_VALUE:
                            mass[num] = ''
                        yum_phone0 = num
                    elif cell == PHONES_LABELS[1]:
                        if mass[num] == ERROR_VALUE:
                            mass[num] = ''
                        yum_phone1 = num
                    elif cell == PHONES_LABELS[2]:
                        if mass[num] == ERROR_VALUE:
                            mass[num] = ''
                        yum_phone2 = num
                    elif mass[num] == ERROR_VALUE:
                        yum = False
                else:
                    mass[num] = ''                # заполняем mass, чтобы его добавить как строку в xlsx

#            yam = 0
#            if len(phone.tel_mob) > 0:
#                yam = int(phone.tel_mob)
#            if len(phone.tel_rod) > 0:
#                yam = yam + int(phone.tel_rod)
#            if len(phone.tel_dom) > 0:
#                yam = yam + int(phone.tel_dom)

            if mass[yum_phone0] == mass[yum_phone1] and mass[yum_phone0] !='':      # стираем дублирующиеся телефоны
                mass[yum_phone1] = ''
            if mass[yum_phone1] == mass[yum_phone2] and mass[yum_phone1] !='':
                mass[yum_phone2] = ''
            if mass[yum_phone0] == mass[yum_phone2] and mass[yum_phone0] !='':
                mass[yum_phone2] = ''

            if mass[yum_phone0] == '' and mass[yum_phone1] == '' and mass[yum_phone2] == '':
                yum = False                                                  # если нет ни одного телефона - ошибка


            if yum and err_from_log.get(num_row + 1) == None:
                ws.append(mass)
#                print(num_row, result_row['ФИО.Фамилия'], result_row['ФИО.Имя'], result_row['ФИО.Отчество'])
            else:
                mass.append(num_row + 1)
                mass.append(err_from_log.get(num_row + 1))
                ws_err.append(mass)
#                print(num_row, result_row['ФИО.Фамилия'], result_row['ФИО.Имя'], result_row['ФИО.Отчество'])
            if num_row % 10000 == 0:                # режем по 10000
                f = self.fname.replace(self.fname.split('/')[-1], '{0:02d}'.format(file_number) + '_' +
                                       self.fname.split('/')[-1])
                wb.save(f)
                wb = Workbook(write_only=True)
                ws = wb.create_sheet('Лист1')
                ws.append(HEAD_RESULT_EXCEL_FILE)  # добавляем первую строку xlsx
                file_number += 1

        f = self.fname.replace(self.fname.split('/')[-1], '{0:02d}'.format(file_number) + '_'+ self.fname.split('/')[-1])
        wb.save(f)
        f = self.fname.replace(self.fname.split('/')[-1], 'err'.format(file_number) + self.fname.split('/')[-1])
        wb_err.save(f)
        dbconn.close()
        if use_log:
            log_file.close()
        if GENERATE_SNILS:
            self.wb_comp.save(self.fname.replace(self.fname.split('/')[-1], 'com'.format(file_number)
                                                 + self.fname.split('/')[-1]))


class BaseClass:

    def __setattr__(self, name, value):
        if isinstance(value, (int, str)):
            self.__dict__[name] = str(value).strip()
        else:
            self.__dict__[name] = value


def normalize(*args):
    result = []
    for arg in args:
        if arg == NULL_VALUE:
            result.append(NEW_NULL_VALUE)
        else:
            result.append(str(arg).strip())
    return result


def normalize_snils(snils):
    snils = str(snils).strip()
    snilsX = ''
    if snils != NULL_VALUE and snils != '' and isinstance(snils, str):
        try:
            for cc in snils:
                if cc in string.digits:
                    snilsX = snilsX+cc
            if len(snilsX) < LEN_SNILS:
                snilsX = '0' * (LEN_SNILS - len(snilsX)) + snilsX
            elif len(snilsX) == LEN_SNILS:
                pass
            else:
                return ERROR_VALUE
            return snilsX
        except TypeError:
            return ERROR_VALUE
    else:
        return ERROR_VALUE


def field2fio(field):
    first_name, second_name, third_name = NEW_NULL_VALUE_FOR_ALL_TEXT, NEW_NULL_VALUE_FOR_ALL_TEXT, \
                                          NEW_NULL_VALUE_FOR_ALL_TEXT
    if len(field) > 0 and field != NULL_VALUE:
        while field.find('  ') > -1:
            field = field.replace('  ', ' ')
        words = field.split()
        for i, word in enumerate(words):
            if i == 0:
                first_name = words[i]
            elif i == 1:
                second_name = words[i]
            elif i == 2:
                third_name = words[i]
            else:
                third_name += ' ' + words[i]
        third_name = third_name.strip()
    return first_name, second_name, third_name



def field2addr(field):
    addr_name, addr_type = '', ''
    if len(field) > 0 and field != NULL_VALUE:
        new_field = ''
        for i, ch in enumerate(field):          # убираем точки и запятые
            if ch == '.' or ch == ',':
                ch = ''
            new_field = new_field + ch
        field = new_field.strip().split(' ')
        TYPES = [REG_TYPES, DISTRICT_TYPES, CITY_TYPES, NP_TYPES, STREET_TYPES]
        for i, word in enumerate(field):
            addr_type_vrem = ''
            for label_group in TYPES:
                for label in label_group:
                    if word.lower() == label.lower():
                        addr_type_vrem = label
            if addr_type_vrem == '':
                addr_name = addr_name + ' ' + word
            else:
                addr_type = addr_type_vrem
    return addr_name, addr_type

#class Gender(BaseClass):
#    def __init__(self, third_name='', gender_field_exists=False, gender=''):
#        self.female_gender_value = female_gender_value
#        self.male_gender_value = male_gender_value
#        self.third_name = str(third_name).strip()
#        self.gender_field_exists = gender_field_exists
#        self.gender = gender.strip()

#    def gender_from_fio(self):
#        if self.third_name == '':
#            return ERROR_VALUE
#        third_name = self.third_name
#        third_name = third_name.split(' ')
#        if len(third_name) == 1:
#            if ''.join(third_name[0][-3:]).lower() == 'вна':
#                gender = '0'
#            elif ''.join(third_name[0][-3:]).lower() == 'вич':
#                gender = '1'
#            else:
#                gender = ERROR_VALUE
#        elif len(third_name) == 2:
#            if third_name[-1].lower() in EAST_GENDER[0]:  # женщина
#                gender = '0'
#            elif third_name[-1].lower() in EAST_GENDER[1]:  # мужчина
#                gender = '1'
#            else:
#                gender = ERROR_VALUE
#        else:
#            gender = ERROR_VALUE
#        return gender

#    def set_gender_value(self, male_value, female_value):
#        self.female_gender_value = female_value.lower()
#        self.male_gender_value = male_value.lower()

#    def get_gender_value(self):
#        return self.female_gender_value, self.male_gender_value

#    def normalize_gender(self):
#        gender = self.gender
#        gender = gender.lower()
#        if gender == self.female_gender_value:
#            return '0'
#        elif gender == self.male_gender_value:
#            return '1'
#        else:
#            return self.gender_from_fio()

#    def get_value(self):
#        if self.gender_field_exists:
#            return self.normalize_gender()
#        else:
#            return self.gender_from_fio()

def normalize_gender(gender, female_gender_value, male_gender_value, gender_length):
    gender = str(gender).strip()
    if gender =='':
        return NEW_NULL_VALUE_FOR_GENDER
    elif len(gender) > 1 and (gender.strip().upper()[:gender_length] != female_gender_value and
                              gender.strip().upper()[:gender_length] != male_gender_value):
        return NEW_NULL_VALUE_FOR_GENDER
    else:
        if gender.strip().upper()[:gender_length] == female_gender_value:
            return '1'
        else:
            return '0'

def normalize_text(tx):
    tx = str(tx).strip()
    if len(tx) <= 1:
        return NEW_NULL_VALUE_FOR_ALL_TEXT
    else:
        return tx

def normalize_date(date):
    date = str(date)
    try:
        result = re.findall(r'\b(\d{4}|\d{2})[\.:-](\d{2})[\.:-](\d{4}|\d{2})\b', date)
        if len(result) > 0:
            if result[0] == NULL_VALUE:
                return NEW_NULL_VALUE_FOR_DATE
            if len(result[0][0]) == 4:
                result[0] = result[0][::-1]
            elif len(result[0][2]) == 2:
                if result[0][2] < 20:
                    result[0][2] = '20' + result[0][2]
                elif result[0][2] > 20:
                    result[0][2] = '19' + result[0][2]
            return '.'.join(result[0])
        else:
            return NEW_NULL_VALUE_FOR_DATE
    except Exception as ee:
        return  NEW_NULL_VALUE_FOR_DATE


# print(normalize_date('01.09.2003'))

# normalize место рождения


class Passport(BaseClass):
    def __init__(self, seriya='', nomer='', date='', who='', cod=''):
        self.seriya = str(seriya).strip()
        self.nomer = str(nomer).strip()
        self.date = normalize_date(date)
        self.who = normalize_text(who)
        self.cod = str(cod).strip()

    def __setattr__(self, name, value):
        if name == 'date':
            self.__dict__[name] = normalize_date(value)
        else:
            if isinstance(value, (int, str)):
                self.__dict__[name] = str(value)
            else:
                self.__dict__[name] = value

    def normalize_seriya(self):
        if self.seriya != NULL_VALUE and self.seriya != '' and isinstance(self.seriya, str):
            try:
                self.seriya = ''.join([char for char in self.seriya if char in string.digits])
                if len(self.seriya) == 3:
                    self.seriya = '0' + self.seriya
                elif len(self.seriya) == 4:
                    pass
                else:
                    return NEW_NULL_VALUE_FOR_SERIYA_PASSPORTA
#                   return ERROR_VALUE
#                self.seriya = self.seriya[:2] + ' ' + self.seriya[2:]         # Тот самый пробел между ## ##
                return self.seriya
            except TypeError:
                return NEW_NULL_VALUE_FOR_SERIYA_PASSPORTA
        else:
            return NEW_NULL_VALUE_FOR_SERIYA_PASSPORTA

    def normalize_nomer(self):
        if self.nomer != NULL_VALUE and self.nomer != '' and isinstance(self.seriya, str):
            try:
                self.nomer = ''.join([char for char in self.nomer if char in string.digits])
                if len(self.nomer) < LEN_PASSPORT_NOMER and len(self.nomer) > 0 :
                    self.nomer = '0' * (LEN_PASSPORT_NOMER - len(self.nomer)) + self.nomer
                elif len(self.nomer) == LEN_PASSPORT_NOMER:
                    pass
                else:
                    return NEW_NULL_VALUE_FOR_NOMER_PASSPORTA
#                    return ERROR_VALUE
                return self.nomer
            except TypeError:
                return NEW_NULL_VALUE_FOR_NOMER_PASSPORTA
        else:
            return NEW_NULL_VALUE_FOR_NOMER_PASSPORTA


    def normalize_who(self):
        if self.who == NULL_VALUE:
            self.who = NEW_NULL_VALUE
        return self.who

    def normalize_cod(self):
        if self.cod != NULL_VALUE and self.cod != '':
            self.cod = ''.join([char for char in self.cod if char in string.digits])
            if len(self.cod) < LEN_PASSPORT_COD and len(self.cod) > 0:
                self.cod = '0' * (LEN_PASSPORT_COD - len(self.cod)) + self.cod
            elif len(self.cod) == LEN_PASSPORT_COD:
                pass
            else:
                return NEW_NULL_VALUE_FOR_COD_PASSPORTA
#                return ERROR_VALUE
            self.cod = self.cod[:3] + '-' + self.cod[3:]
            return self.cod
        else:
            return NEW_NULL_VALUE_FOR_COD_PASSPORTA

    def get_values(self):
        return self.normalize_seriya(), self.normalize_nomer(), self.date, self.normalize_who(), self.normalize_cod()


def normalize_index(index):
    index = str(index).strip()
    if index != NULL_VALUE and index != '' and isinstance(index, str):
        try:
            index = ''.join([char for char in index if char in string.digits])
            if len(index) < LEN_INDEX_NOMER:
                index = '0' * (LEN_INDEX_NOMER - len(index)) + index
            elif len(index) == LEN_INDEX_NOMER:
                pass
            else:
                return NEW_NULL_VALUE_FOR_INDEX
#                return ERROR_VALUE
            return index
        except TypeError:
            return NEW_NULL_VALUE_FOR_INDEX
    else:
        return NEW_NULL_VALUE_FOR_INDEX

def normalize_home(tx):
        tx = str(tx).strip()
        numbers = True
        for i in range(len(tx)):
            if tx[i] not in string.digits:
                numbers = False
        if len(tx) < 1:
            return tx
        elif len(tx) > 10:
            return NEW_NULL_VALUE_FOR_HOME
        elif numbers:
            if int(tx) > 1500:
                return NEW_NULL_VALUE_FOR_HOME
            else:
                return tx
        else:
            return tx


class FullAdress(BaseClass):
    def __init__(self, field='', tip='стандартный'):
    #def __init__(self, field='', tip='стандартный'):
        self.field = str(field)
        self.field_home = ''
        self.homes = []
        self.index = []
        self.tip = tip
        self.full_adress = []
        self.FULL_ADRESS_DICT = {}
        for label in FULL_ADRESS_LABELS:
            self.FULL_ADRESS_DICT[label] = ''
        self.iter_types = [DISTRICT_TYPES, CITY_TYPES, NP_TYPES, STREET_TYPES, HOUSE_CUT_NAME, CORPUS_CUT_NAME, APARTMENT_CUT_NAME]

    def normalize_adress(self):
        if len(self.field) != 0 and self.field != NULL_VALUE:
            self.field = self.field.lower()
            values = self.field.split(SPLIT_FIELD) # разделили на массив разделителем SPLIT_FIELD
            for i, word in enumerate(values):
                n = []
                word = word.strip()
                if i == 0:
                    n = [char for char in word if char in digits]
                    if len(n) != 6:
                        return NEW_NULL_VALUE_FOR_INDEX
                    self.FULL_ADRESS_DICT[FULL_ADRESS_LABELS[0]] = ''.join(n)
                    continue
                elif i == 1:
                    self.FULL_ADRESS_DICT[FULL_ADRESS_LABELS[1]] = ' '.join(word.split(' ')[:-1])
                    self.FULL_ADRESS_DICT[FULL_ADRESS_LABELS[2]] = word.split(' ')[-1]
                    continue
                else:
                    for j, types in enumerate(self.iter_types):
                        if j < 4:
                            if word.split(' ')[-1] in types:
                                self.FULL_ADRESS_DICT[FULL_ADRESS_LABELS[3 + 2 * j]] = ' '.join(
                                    word.split(' ')[:-1])
                                self.FULL_ADRESS_DICT[FULL_ADRESS_LABELS[3 + 2 * j + 1]] = word.split(' ')[-1]
                        elif j >= 4:
                            for type in types:
                                if word.find(type) != (-1):
                                    word = word.replace(type, '').replace('.', '')
                                    self.FULL_ADRESS_DICT[FULL_ADRESS_LABELS[11 + j - 4]] = word
            return self.FULL_ADRESS_DICT
        else:
            if self.field == NULL_VALUE:
                return NEW_NULL_VALUE
            else:
                return NEW_NULL_VALUE
                #                return ERROR_VALUE

    def create_output_list(self):                   # Создаем список вывода
        if self.field != '':
            FULL_ADRESS_DICT = self.normalize_adress()
        for label in FULL_ADRESS_LABELS:
            self.full_adress.append(self.FULL_ADRESS_DICT[label].upper())
        return self.full_adress

    def cut_adress(self):
        self.field = self.field.lower()
        # Заменить разделители на пробелы, схлопнуть двойные пробелы в одинарные
        field_cut = self.field
        for cut in SPLIT_FIELDS:
            field_cut = field_cut.replace(cut, ' ')
        while field_cut.find('  ') > -1:
            field_cut = field_cut.replace('  ', ' ')
        words = field_cut.split()
        # разделяем дома и деревни
        field_cut = ''
        for i, word in enumerate(words):
            if i > len(words) - 3:
                break
            if word == 'д':
                is_willage = True
                for char in words[i + 1]:
                    if char in digits:
                        is_willage = False
                        break
                if is_willage:
                    words[i] = 'дер'
        field_cut = ' '.join(words)
        self.field = field_cut.strip()
        field_cut = ''
        breaks = {}
        breaks_len = {}
        breaks_name = {}
        for adress_type in ADRESS_TYPES:
            for left in SPLIT_FIELDS:
                for right in SPLIT_FIELDS:
                    if len(self.field.split(left + adress_type + right)) > 1:
                        try:
                            tek_pos = 0
                            for i, tek_part in enumerate(self.field.split(left + adress_type + right)):
                                if i == len(self.field.split(left + adress_type + right)) - 1:
                                    break
                                tek_pos += len(tek_part) + len(left + adress_type + right)
                                breaks[ADRESS_TYPES[adress_type]].append(tek_pos)
                        except KeyError:
                            breaks[ADRESS_TYPES[adress_type]] = []
                            tek_pos = 0
                            for i, tek_part in enumerate(self.field.split(left + adress_type + right)):
                                if i == len(self.field.split(left + adress_type + right)) - 1:
                                    break
                                tek_pos += len(tek_part) + len(left + adress_type + right)
                                breaks[ADRESS_TYPES[adress_type]].append(tek_pos)
                        try:
                            for i, tek_part in enumerate(self.field.split(left + adress_type + right)):
                                if i == len(self.field.split(left + adress_type + right)) - 1:
                                    break
                                breaks_name[ADRESS_TYPES[adress_type]].append(adress_type)
                        except KeyError:
                            breaks_name[ADRESS_TYPES[adress_type]] = []
                            for i, tek_part in enumerate(self.field.split(left + adress_type + right)):
                                if i == len(self.field.split(left + adress_type + right)) - 1:
                                    break
                                breaks_name[ADRESS_TYPES[adress_type]].append(adress_type)

        digits_count = 0
        digits_pos = 0
        index_pos = -1
        for i, char in enumerate(self.field):
            if char in digits:
                if i - digits_pos > 1 or digits_count > 6:
                    digits_count = 1
                else:
                    digits_count += 1
                digits_pos = i
                if digits_count == 6:
                    index_pos = i - 5
                    try:
                        breaks[0].append(index_pos)
                    except KeyError:
                        breaks[0] = []
                        breaks[0].append(index_pos)
                    try:
                        breaks_len[0].append(6)
                    except KeyError:
                        breaks_len[0] = []
                        breaks_len[0].append(6)
                    try:
                        breaks_name[0].append('')
                    except KeyError:
                        breaks_name[0] = []
                        breaks_name[0].append('')

        # сортируем по значениям словаря
        breaks_sorted = OrderedDict(sorted(breaks.items(), key=lambda t: t[1]))

        # Отсечь дом-корпус-квартиру
        cutted_begin = []
        cutted_end = []
        home_cut = 0
        home_pass = 0
        tek_home = ''
        last_i = 0
        for i in range(5):
            #last_break_sorted = list(breaks_sorted.keys())[0]
            for break_sorted in breaks_sorted:
                if break_sorted == 0 and len(breaks_sorted[break_sorted]) > i: # индекс
                    try:
                        if home_cut: # дом-корпус-квартира (окончание)
                            cutted_end.append(breaks_sorted[break_sorted][i])
                            tek_home += ', ' + self.field[home_pass:breaks_sorted[break_sorted][i]].strip()
                            home_cut = 0
                            self.homes.append(tek_home)
                            tek_home = ''
                            #self.homes.append(self.field[cutted_begin[len(cutted_begin) - 1]:
                            #                             cutted_end[len(cutted_end) - 1]].strip())
                        cutted_begin.append(breaks_sorted[break_sorted][i])
                        cutted_end.append(breaks_sorted[break_sorted][i] + breaks_len[break_sorted][i])
                        self.index.append(self.field[breaks_sorted[break_sorted][i]:breaks_sorted[break_sorted][i] +
                                                                                    breaks_len[break_sorted][i]])
                        last_break_sorted = break_sorted
                    except IndexError:
                        pass
                elif break_sorted == 11 and len(breaks_sorted[break_sorted]) > i: # дом-корпус-квартира (начало)
                    try:
                        cutted_begin.append(breaks_sorted[break_sorted][i] - len(breaks_name[break_sorted][i]) - 1)
                        home_cut = breaks_sorted[break_sorted][i]
                        home_pass = breaks_sorted[break_sorted][i] - len(breaks_name[break_sorted][i]) - 1
                        last_break_sorted = break_sorted
                    except IndexError:
                        pass
                elif break_sorted > 11 and len(breaks_sorted[break_sorted]) > i:
                    if home_cut:
                        try:
                            tek_home += ', ' + self.field[home_pass:breaks_sorted[break_sorted][i]].strip(
                                               )[:-len(breaks_name[break_sorted][i])].strip()
                            home_pass = breaks_sorted[break_sorted][i] - len(breaks_name[break_sorted][i]) - 1
                        except IndexError:
                            pass
                    last_break_sorted = break_sorted
                elif break_sorted < 11 and len(breaks_sorted[break_sorted]) > i: # дом-корпус-квартира (окончание)
                    if home_cut:
                        try:
                            cutted_end.append(breaks_sorted[last_break_sorted][last_i] + len(tek_part))
                            tek_home += ', ' + self.field[home_pass:breaks_sorted[last_break_sorted][last_i] + len(tek_part)].strip()
                            home_cut = 0
                            self.homes.append(tek_home)
                            tek_home = ''
                        except IndexError:
                            pass
                    last_break_sorted = break_sorted
            last_i = i
        if len(cutted_begin) != len(cutted_end):
            cutted_end.append(len(self.field))
            tek_home += ', ' + self.field[home_pass:].strip()
            home_cut = 0
            self.homes.append(tek_home)

            #self.homes.append(self.field[cutted_begin[len(cutted_begin) - 1]:cutted_end[len(cutted_end) - 1]].strip())

        # удалить индекс, дом-корпус-квартиру
        field_cut = ''
        if len(cutted_begin):
            field_cut = self.field[:cutted_begin[0]]
            for i, cutted in enumerate(cutted_begin):
                if i:
                    field_cut += self.field[cutted_end[i-1]:cutted]

        # Заменить разделители на пробелы, схлопнуть двойные пробелы в одинарные
        for cut in SPLIT_FIELDS:
            field_cut = field_cut.replace(cut, ' ')
        while field_cut.find('  ') > -1:
            field_cut = field_cut.replace('  ', ' ')

        # Удалить типы объектов
        for adress_type in ADRESS_TYPES:
            field_cut = field_cut.replace(' ' + adress_type + ' ',' ')
        for adress_type in ALL_CUT_NAMES:
            field_cut = field_cut.replace(' ' + adress_type + ' ', ' ')
        while field_cut.find('  ') > -1:
            field_cut = field_cut.replace('  ', ' ')

        # Удалить дублирующиеся слова
        words = field_cut.split()
        field_cut = ''
        for word in words:
            if word not in field_cut.split():
                field_cut += word + ' '
        self.field_home = field_cut

    def get_values(self):
        if self.tip == 'стандартный':
            # Когда адрес 414000, г. Астрахань, ул. Такая, д. Т...
            output_list = []
            for elem in self.create_output_list():
                output_list.append(elem.strip())
            return output_list
        elif self.tip == 'перемешаный':
            # Когда все поля по раздельности и перемешаны...
            output_list = []

            if len(self.field) != 0 and self.field != NULL_VALUE:
                self.field = self.field.lower()
                values = self.field.split(SPLIT_FIELD)
                for i, nn in enumerate(ORDER_FIELD):
                    if nn < len(values):
                        output_list.append(values[nn])
                return output_list
            else:
                if self.field == NULL_VALUE:
                    return NEW_NULL_VALUE_FOR_ADDRESS
                else:
                    return NEW_NULL_VALUE_FOR_ADDRESS
        elif self.tip == 'КЛАДР':
            # Заменить разделители на пробелы, схлопнуть двойные пробелы в одинарные
            if self.field == '':
                return NEW_NULL_VALUE_FOR_ADDRESS
            field_cut = self.field
            for cut in SPLIT_FIELDS:
                field_cut = field_cut.replace(cut, ' ')
            while field_cut.find('  ') > -1:
                field_cut = field_cut.replace('  ', ' ')
            self.field = field_cut
            all_homes = ''
            output_list = []
            breaked_address = False
            if len(self.field) > 0 and self.field != NULL_VALUE:
                self.cut_adress()
                if len(self.homes):
                    base_home = self.homes[0]
                    all_homes = self.homes[0]
                    for home in self.homes:
                        if home != base_home:
                            all_homes += home
                            breaked_address = True
                if len(self.index):
                    base_index = self.index[0]
                    all_index = self.index[0]
                    for index in self.index:
                        if index != base_index:
                            all_index += ', ' + index
                            breaked_address = True
                else:
                    all_index = '111111'
                try:
                    res = requests.get('http://127.0.0.1:23332/find/' + self.field_home)
                except Exception as e:
                    print('Сервис адресов не запущен')
                    sys.exit()
                if res.status_code == 200:
                    try:
                        ajson = json.loads(bytes.decode(res.content))
                        self.field = all_index + ', ' + ajson[0]['text'] + all_homes
                    except KeyError:
                        if len(self.index) > 1:
                            return ['', '', '', '', '', '', '', '', all_index + self.field_home + all_homes, '', '', '',
                                    '']
                        else:
                            return [all_index,'','','','','','','', self.field_home + all_homes,'','','','']
                    except IndexError:
                        if len(self.index) > 1:
                            return ['', '', '', '', '', '', '', '', all_index + self.field_home + all_homes, '', '', '',
                                    '']
                        else:
                            return [all_index,'','','','','','','', self.field_home + all_homes,'','','','']
                    address_dict = {}
                    address_words = self.field.split(',')
                    # разделяем дома и деревни
                    field_cut = ''
                    for i, word in enumerate(address_words):
                        if word.find(' д ') > -1:
                            is_willage = True
                            for char in address_words[i].replace(' д ',''):
                                if char in digits:
                                    is_willage = False
                                    break
                            if is_willage:
                                address_words[i] = address_words[i].replace(' д ',' дер ')
                        field_cut += address_words[i] + ', '
                    address_words = field_cut.split(',')
                    for address_word in address_words:
                        for address_type in ADRESS_TYPES:
                            if address_word.find(' ' + address_type + ' ') > -1:
                                address_dict[ADRESS_TYPES[address_type]] = address_word.replace(' ' + address_type +
                                                                                                        ' ', '').strip()
                                if ADRESS_TYPES[address_type] and ADRESS_TYPES[address_type] < 10:
                                    address_dict[ADRESS_TYPES[address_type] + 1] = address_type
                    address_words = []
                    for i, address_word in enumerate(FULL_ADRESS_LABELS):
                        if not i:
                            address_words.append(all_index)
                            continue
                        try:
                            address_words.append(address_dict[i])
                        except KeyError:
                            address_words.append('')
                    return address_words
                else:
                    if len(self.index) > 1:
                        return ['', '', '', '', '', '', '', '', all_index + self.field_home + all_homes, '', '', '', '']
                    else:
                        return [all_index, '', '', '', '', '', '', '', self.field_home + all_homes, '', '', '', '']


class Phone(BaseClass):
    def __init__(self, tel_mob='', tel_rod='', tel_dom=''):
        self.tel_mob = str(tel_mob).strip()
        self.tel_rod = str(tel_rod).strip()
        self.tel_dom = str(tel_dom).strip()

    def normalize_tel_number(self, tel):
        tel = tel.strip()
        if tel == '' or tel == NULL_VALUE:
            return ERROR_VALUE
        tel = str(tel).strip()
        tel = ''.join([char for char in tel if char in string.digits])
        if len(tel) == 11:
            if tel[0] in ['7', '8', '9']:
                tel = '7' + tel[1:]
            else:
                return ERROR_VALUE
        elif len(tel) == 10:
            tel = '7' + tel
        else:
            return ERROR_VALUE
        return tel

    def poryadoc(self, *tels):
        tels = sorted(tels)
#        tels.reverse()
        return list(tels)

    def get_values(self):
        self.tel_mob = self.normalize_tel_number(self.tel_mob)
        self.tel_rod = self.normalize_tel_number(self.tel_rod)
        self.tel_dom = self.normalize_tel_number(self.tel_dom)
#        if self.tel_rod == self.tel_mob:
#            self.tel_rod = ''
#        self.tel_dom = self.normalize_tel_number(self.tel_dom)
#        if self.tel_dom == self.tel_mob or self.tel_dom == self.tel_rod:
#            self.tel_rod = ''
        return self.poryadoc(self.tel_mob, self.tel_rod, self.tel_dom)

# p = Phone()
# p.tel_dom = 89040964007
# p.tel_rod = 89257349331
# p.tel_mob = 'dd'
# print(p.get_values())

def intl(a):               # белиберду в цифры или 0
    try:
        if a != None:
            a = str(a).strip()
            if  a != '':
                a = ''.join([char for char in a if char in string.digits])
                if len(a) > 0:
                    return int(a)
                else:
                    return 0
        return 0
    except TypeError:
        return 0

def field2sernum(field):
    if len(field) > 0 and field != NULL_VALUE and l(field) > 999999 and len('{0:010d}'.format(l(field))) == 10:
        new_field = '{0:010d}'.format(l(field))
        seria = new_field[:4]
        number = new_field[4:]
        return seria, number
    else:
        return '1111', '111111'

def normalize_seria(field):
    if len(field) > 0 and field != NULL_VALUE and l(field) > 0 and len('{0:04d}'.format(l(field))) == 4:
        return '{0:04d}'.format(l(field))
    else:
        return '1111'


